import openpyxl as xl
from openpyxl import formatting
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.styles.numbers as format

CUSTOM_FORMAT_CURRENCY_ONE = '_($* #,##0.0_);[Red]_($* (#,##0.0);_($* "-"??_)'
CUSTOM_FORMAT_CURRENCY_TWO = '_($* #,##0.00_);[Red]_($* (#,##0.00);_($* "-"??_)'
letters = ' ABCDEFGHIJKLMNOPQRSTUVWXYZ'
titles = ["Operating Income", "Depreciation Adjustment", "Depreciation", "CAPEX", "Growth CAPEX", "Option Expense", "Interest Earned on Cash",
          "Cash", "Interest Rate", "Pretax Earnings", "Tax Rate", "Taxes", "Earnings", "Earnings Power Value", "Cash", "Debt",
          "Total EV in Equity", "Shares Outstanding", "EPV/share", "Current Share Price"]

depreciation_adjustment = ["Premises and Equipment", "Current Year Revenue", "Prior Year Revenue", "Change in Revenue", 
                           "Depreciation and Amortization", "CAPEX", "Growth CAPEX", "Zero-growth CAPEX", "Depreciation Adjustment"]

boldFont = Font(name='Arial',size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
textFont  = Font(name='Arial',size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
redTextFont = Font(name='Arial',size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='ff0000')

greyFill = PatternFill(fill_type="solid", start_color='999997', end_color='999997')
yellowFill = PatternFill(fill_type="solid", start_color='ffff01', end_color='ffff01')
orangeFill = PatternFill(fill_type="solid", start_color='fde599', end_color='fde599')

epvFill = PatternFill(fill_type="solid", start_color="c6e6c1", end_color="c6e6c1")
greenFill = PatternFill(fill_type="solid", start_color='7eab76', end_color='7eab76')
notesFill = PatternFill(fill_type="solid", start_color='fdd966', end_color='fdd966')
greeNotesFill = PatternFill(fill_type="solid", start_color='b6d7a8', end_color='b6d7a8')

depAdjFill = PatternFill(fill_type="solid", start_color="fef2cc", end_color="fef2cc")
depAdjDataFill = PatternFill(fill_type="solid", start_color="cfe2f3", end_color="cfe2f3")

titleFill = PatternFill(fill_type="solid", start_color='ebebeb', end_color='ebebeb')
thinBorder = Side(style="thin", color="000000")
thickBorder = Side(style="thick", color="000000")
noBorder = Side(style="none")

def cell_address(col, row):
    return f"{letters[col]}{row}"

def format_for_sign(value):
    return textFont if value >= 0 else redTextFont

def EPV_titles(wb_EPV, row, col):
    wb_EPV.column_dimensions['B'].width = 30
    wb_EPV.cell(row=2, column=2, value="")
    cell = wb_EPV["B2"]
    cell.fill = greyFill
    cell.border = Border(left=thickBorder, top=thickBorder, right=noBorder, bottom=noBorder)

    for i, value in enumerate(titles):
        wb_EPV.cell(row=row, column=col, value=value)
        cell = wb_EPV[f"B{row}"]
        cell.font = boldFont
        cell.fill = titleFill
        
        if (row == 4 or row == 7 or row == 20):
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thinBorder)
        elif (row == 22):
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thickBorder)
        else:
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        row += 1
    
def dep_adj_titles(wb_EPV, row, col):
    for i, value in enumerate(depreciation_adjustment):
        wb_EPV.cell(row=row, column=col, value=value)
        cell = wb_EPV[f"B{row}"]
        cell.font = textFont
        cell.fill = depAdjFill
        
        if (row == 24):
            cell.border = Border(left=thickBorder, top=thickBorder, right=noBorder, bottom=noBorder)
        elif (row == 31):
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thinBorder)
        elif (row == 32):
            cell.font = boldFont
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thickBorder)
        else:
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        row += 1

def fill_EPV_data(wb_EPV, row, col, epv_info):
    col = 3
    for y, date in enumerate(epv_info):
        wb_EPV.column_dimensions[f"{letters[col]}"].width = 25
        row = 2
        year = date.split('-')[0]
        wb_EPV.cell(row=row, column=col, value=year)
        cell = wb_EPV[f"{letters[col]}2"]
        cell.fill = greenFill
        cell.alignment = Alignment(horizontal="center")
        cell.font = boldFont
        cell.border = Border(left=thinBorder, top=thickBorder, right=thinBorder, bottom=noBorder)
        row += 1
        
        for i, epv_value in enumerate(titles):
            value = epv_info[date][epv_value] if epv_value in epv_info[date] else None
            # Debt needs to be negative
            if epv_value == "Debt":
                value = -abs(value)

            wb_EPV.cell(row=row, column=col, value=value)
            cell = wb_EPV[f"{letters[col]}{row}"]
            #cell.font = titleFont
            cell.fill = epvFill
            
            # Depreciation Adjustment
            if row == 4:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{32}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=thinBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Depreciation
            elif row == 5:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{28}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # CAPEX
            elif row == 6:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{29}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)      
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Growth CAPEX
            elif row == 7:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{30}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=thinBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Interest Earned on Cash
            elif row == 9:
                wb_EPV.cell(row=row, column=col, value=f"=-{letters[col]}{10}*{letters[col]}{11}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            elif row == 11:
                int_year = int(year)
                if int_year >= 2023:
                    year_val = 0.06
                elif int_year == 2022:
                    year_val = 0.04
                else:
                    year_val = 0.02
                wb_EPV.cell(row=row, column=col, value=year_val)
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = format.FORMAT_PERCENTAGE_00
            # Pretax Earnings
            elif row == 12:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{3}+{letters[col]}{4}+{letters[col]}{9}-{letters[col]}{8}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Tax Rate
            elif row == 13:
                wb_EPV.cell(row=row, column=col, value=0.21)
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = format.FORMAT_PERCENTAGE_00
            # Taxes
            elif row == 14:
                wb_EPV.cell(row=row, column=col, value=f"=-{letters[col]}{12}*{letters[col]}{13}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Earnings
            elif row == 15:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{12}+{letters[col]}{14}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Earnings Power Value
            elif row == 16:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{15}/0.1")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Cash
            elif row == 17:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{10}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Total EV in Equity
            elif row == 19:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{16}+{letters[col]}{17}+{letters[col]}{18}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Shares Outstanding
            elif row == 20:
                wb_EPV.cell(row=row, column=col, value=8650)
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=thinBorder)
                cell.number_format = format.FORMAT_NUMBER
            # EPV/Share
            elif row == 21:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{19}/{letters[col]}{20}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
                cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
                cell.fill = yellowFill
                cell.font = boldFont
            # Current Share Price
            elif (row == 22):
                wb_EPV.cell(row=row, column=col, value=130.00)
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=thickBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
                cell.fill = yellowFill
                cell.font = boldFont
            else:
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            
            #cell.font = format_for_sign(value)
            row += 1 
        col += 1

def fill_dep_adj_data(wb_EPV, row, col, epv_info):
    col = 3
    for y, date in enumerate(epv_info):
        row = 24
        for i, epv_value in enumerate(depreciation_adjustment):
            value = epv_info[date][epv_value] if epv_value in epv_info[date] else None
            # CAPEX needs to always positive
            if epv_value == "CAPEX":
                value = abs(value)
            wb_EPV.cell(row=row, column=col, value=value)
            cell = wb_EPV[f"{letters[col]}{row}"]
            cell.font = textFont
            cell.fill = depAdjDataFill
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            
            # Premises and Equipment
            if (row == 24):
                cell.border = Border(left=thinBorder, top=thickBorder, right=thinBorder, bottom=noBorder)
            # Prior Year Revenue
            elif row  == 26:
                if letters[col] == 'C':
                    wb_EPV.cell(row=row, column=col, value=None)
                else:
                    wb_EPV.cell(row=row, column=col, value=f"={letters[col-1]}{25}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
            # Change in Revenue
            elif row == 27:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{25}-{letters[col]}{26}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)

            # Growth CAPEX:
            elif row == 30:
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{24}/{letters[col]}{25}*{letters[col]}{27}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
            # Zero-growth CAPEX
            elif (row == 31):
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{29}-{letters[col]}{30}")
                cell.border = Border(left=thinBorder, top=noBorder, right=thinBorder, bottom=thinBorder)
            # Depreciation Adjustment
            elif (row == 32):
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{28}-{letters[col]}{31}")
                cell.font = boldFont
                cell.border = Border(left=thinBorder, top=noBorder, right=thinBorder, bottom=thickBorder)
            else:
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            row += 1
        col += 1

def fill_notes(wb_EPV, row, col):
    cell = wb_EPV[f"{letters[col]}{row}"]
    wb_EPV.cell(row=row, column=col, value="Notes")
    wb_EPV.column_dimensions[f"{letters[col]}"].width = 15
    cell.font = boldFont
    cell.alignment = Alignment(horizontal="center")
    cell.fill = notesFill
    cell.border = Border(left=thinBorder, top=thickBorder, right=thickBorder, bottom=noBorder)

    # EPV notes
    row = 3
    for _ in range(3, 23):
        wb_EPV.cell(row=row, column=col, value=None)
        cell = wb_EPV[f"{letters[col]}{row}"]
        cell.fill = orangeFill  

        if (row == 22):
            cell.border = Border(left=thinBorder, top=noBorder, right=thickBorder, bottom=thickBorder)
        else:
            cell.border = Border(left=thinBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        row += 1
    
    # DA Notes
    row = 24
    for _ in range(24, 33):
        wb_EPV.cell(row=row, column=col, value=None)
        cell = wb_EPV[f"{letters[col]}{row}"]
        cell.fill = greeNotesFill
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=thinBorder, top=noBorder, right=thickBorder, bottom=noBorder)

        if (row == 24):
            wb_EPV.cell(row=row, column=col, value="(a)")
            cell.border = Border(left=thinBorder, top=thickBorder, right=thickBorder, bottom=noBorder)
        elif (row == 25):
            wb_EPV.cell(row=row, column=col, value="(b)")
        elif (row == 26):
            wb_EPV.cell(row=row, column=col, value="(c)")
        elif (row == 27):
            wb_EPV.cell(row=row, column=col, value="(d) = (b) - (c)")
        elif (row == 28):
            wb_EPV.cell(row=row, column=col, value="(e)")
        elif (row == 29):
            wb_EPV.cell(row=row, column=col, value="(f)")
        elif (row == 30):
            wb_EPV.cell(row=row, column=col, value="(g) = (a)/(b) x (d)")
        elif (row == 31):
            cell.border = Border(left=thinBorder, top=noBorder, right=thickBorder, bottom=thinBorder)
            wb_EPV.cell(row=row, column=col, value="(h) = (f) - (g)")
        elif row == 32:
            cell.border = Border(left=thinBorder, top=thinBorder, right=thickBorder, bottom=thickBorder)
            wb_EPV.cell(row=row, column=col, value="(i) = (e) - (h)")
        row += 1
    

def fill_epv(wb_EPV, epv_info):
    row, col = 3, 2
    EPV_titles(wb_EPV, row, col)
    row = 24
    dep_adj_titles(wb_EPV, row, col)
    row, col = 2, 3
    fill_EPV_data(wb_EPV, row, col, epv_info)
    row = 24
    fill_dep_adj_data(wb_EPV, row, col, epv_info)
    row, col = 2, 3+len(epv_info)
    fill_notes(wb_EPV, row, col)

if __name__ == "__main__":
    ticker = "DOMI"
    path = f"./EPV/{ticker}.xlsx"	

    wb = xl.Workbook()
    wb.create_sheet("EPV")

    row, col = 3, 2
    wb_EPV = wb["EPV"]
    EPV_titles(wb_EPV, row, col)
    row = 24
    dep_adj_titles(wb_EPV, row, col)
    row, col = 3, 3
    fill_EPV_data(wb_EPV, row, col)
    row = 24
    fill_dep_adj_data(wb_EPV, row, col)
    wb.save(path)