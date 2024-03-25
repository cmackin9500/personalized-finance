import os
from bs4 import BeautifulSoup
import sys
from dataclasses import dataclass
from typing import List
import pandas as pd
import json
import openpyxl as xl

from files import read_forms_from_dir, find_latest_form_dir, find_all_form_dir
from xbrl_parse import get_fs_fields, get_disclosure_fields
from edgar_retrieve import get_company_CIK, get_forms_of_type_xbrl, save_all_facts, save_all_forms
from html_parse import html_to_facts
from html_process import derived_fs_table, assign_HTMLFact_to_XBRLNode
from util.write_to_csv import *
from util.file_management import read_file
from excel_model.epv import *
from excel_model.cover import *
from excel_model.gv import *
from excel_model.wacc import *

def get_fs_list(fs_fields, mag, fs):
	div = 1000
	if mag == 't':
		div = 1000
	elif mag == 'm':
		div = 1000000
	elif div == 'n':
		div = 1

	fs_tag = ""
	if fs == 'bs': fs_tag = "us-gaap:Assets"

	elif fs == 'is': 
		if "us-gaap:NetIncomeLoss" in fs_fields:
			fs_tag = "us-gaap:NetIncomeLoss"
		elif "us-gaap:ProfitLoss" in fs_fields:
			fs_tag = "us-gaap:ProfitLoss"
		elif "us-gaap:OperatingIncomeLoss" in fs_fields:
			fs_tag = "us-gaap:OperatingIncomeLoss"

	elif fs == 'cf':
		if "us-gaap:NetIncomeLoss" in fs_fields:
			fs_tag = "us-gaap:NetIncomeLoss"
		elif "us-gaap:ProfitLoss" in fs_fields:
			fs_tag = "us-gaap:ProfitLoss"
		elif "us-gaap:OperatingIncomeLoss" in fs_fields:
			fs_tag = "us-gaap:OperatingIncomeLoss"
	
	assert fs_tag != "", "The base tag is not found in the financial statement. Look for another tag that'll be in it."
	date = fs_fields[fs_tag].date[0]

	fs_list = []
	for key in fs_fields:
		if fs_fields[key].val is None:
			continue
		text = fs_fields[key].text[0] if fs_fields[key].text else ''
		for i,tag in enumerate(fs_fields[key].text):
			if fs_fields[key].tag == "us-gaap:EarningsPerShareBasic" or fs_fields[key].tag == "us-gaap:EarningsPerShareDiluted":
				d = {'Tag': fs_fields[key].tag, 'Text':text, date: fs_fields[key].val[i]}	
			else:
				d = {'Tag': fs_fields[key].tag, 'Text':text, date: fs_fields[key].val[i]/div}
			fs_list.append(d)
	return fs_list

def get_all_dates(data, period, currency):
	dates1 = []
	cash = 'CashAndCashEquivalentsAtCarryingValue'
	if cash in data: 
		if currency in data[cash]['units']:
			d = list(data[cash]['units'][currency])
			for year in d:
				if year['form'] == period:
					if year['end'] not in dates1: dates1.append(year['end'])
	dates1.sort()

	dates2 = []
	cash = 'CashAndDueFromBanks'
	if cash in data: 
		if currency in data[cash]['units']:
			d = list(data[cash]['units'][currency])
			for year in d:
				if year['form'] == period:
					if year['end'] not in dates1: dates1.append(year['end'])
	dates1.sort()
	dates2.sort()
	return dates1 if len(dates1) > len(dates2) else dates2
	
def get_tags(destination, period):
	div = 1000
	if mag == 't':
		div = 1000
	elif mag == 'm':
		div = 1000000
	elif div == 'n':
		div = 1

	with open(destination, 'r') as f:
		data = f.read()
	facts_json = json.loads(data)['facts']['us-gaap']
	with open("./tags/facts_tags.json", 'r') as f:
		data = f.read()
	USGAAP_json = json.loads(data)

	dates = get_all_dates(facts_json, period,'USD')
	# Special cases where companies like CP has 'CAD' as currecny
	if (dates == []):
		dates = get_all_dates(facts_json, period,'CAD')

	df = pd.DataFrame(columns = dates)

	for category in USGAAP_json:
		for usgaap_tag in USGAAP_json[category]:
			if usgaap_tag in facts_json: 
				rowData = {y:0 for y in dates}
				if 'USD' in facts_json[usgaap_tag]['units']:
					d = list(facts_json[usgaap_tag]['units']['USD'])
					for year in d:
						if year['form'] == period and year['end'] in dates:
							if category == "EPS" or category == "Dividend":
								rowData[year['end']] = year['val']
							else:
								rowData[year['end']] = year['val']/div
				appendRow = [rowData[key] for key in rowData]
				df.loc[usgaap_tag] = appendRow
	return df

def epv(epv_path, json_path, appeared_tags):
	with open(json_path, 'r') as f:
		data = f.read()
	data = json.loads(data)['facts']['us-gaap']
	usd = (list(data)[0])
	if 'USD' not in data[usd]['units']:
		return pd.DataFrame()
	
	with open(epv_path, 'r') as f:
		all_tags = f.read()
	all_tags = json.loads(all_tags)

	dates = get_all_dates(data, '10-K', 'USD')

	df = pd.DataFrame(columns = dates)
	for tags in all_tags:
		all_usgaap_rowData = []
		rowData = {y:0 for y in dates}
		for tag in all_tags[tags]:
			#if tag in data and tag in appeared_tags: 
			if tag in data: 
				if 'USD' in data[tag]['units']:
					d = list(data[tag]['units']['USD'])
					for year in d:
						if year['form'] == '10-K' and year['end'] in dates:
							#TODO: I need to handle fields that adds up to one category. For example debt usually has multiple that adds up to one.
							if rowData[year['end']] != 0: 
								continue 
							rowData[year['end']] += year['val']/div

							# Adding this part to also export all of the elements that sums up to the calulation parameter for EPV
							if {tag: rowData} not in all_usgaap_rowData:
								all_usgaap_rowData.append({tag: rowData})

		# Append the new category in EPV
		appendRow = [rowData[key] for key in rowData]			
		df.loc[tags] = appendRow

		# Also show all the us-gaap tags that calculates up to it
		for usgaap_rowData in all_usgaap_rowData:
			for usgaap_tag in usgaap_rowData:
				rowData = usgaap_rowData[usgaap_tag]
				appendRow = [rowData[key] for key in rowData]			
				df.loc[usgaap_tag] = appendRow
		
	#df.to_csv("ticker.csv")
	return df
					
def fs_process_from_cfiles(cfiles, fs, get_both_dates = False):
	# Get the Balance Sheet information from the most recent 10-K/10-Q.
	fs_fields = get_fs_fields(ticker, fs, cfiles)
	all_tables = html_to_facts(cfiles.html, cfiles.htm_xml, fs_fields)
	fs_table_from_html = derived_fs_table(all_tables, fs_fields)

	index = 0
	if get_both_dates:
		assert len(fs_table_from_html[0]) > 0, "Facts retrieved for the Financial Statement is empty."
		date1 = fs_table_from_html[0][0].date
		date2 = fs_table_from_html[0][1].date
		if date1 > date2: index = 1
	
	FS = assign_HTMLFact_to_XBRLNode(fs_fields, fs_table_from_html, index)
	
	if get_both_dates:
		if index == 1: index = 0
		else: index = 1
		FS = assign_HTMLFact_to_XBRLNode(fs_fields, fs_table_from_html, index)
	return FS

def df_input_fs(FS, div=1000):
	financial_statement = []
	for fact in FS:
		f = FS[fact]
		if f.val is None or f.date is None: continue
		val_date_dict = {f.date[i]: f.val[i]/div for i in range(len(f.date))}
		push_fact = {"Tag": f.text[0]}
		for key in val_date_dict:
			push_fact[key] = val_date_dict[key]
		financial_statement.append(push_fact)
	return financial_statement

def populate_fs_df(fs_list, all_fs_info):
	if all_fs_info == []: 
		all_fs_info = fs_list

	# Loop through the new list and add it to the right place.
	# First, we check of the tag is already added. If so, we just add the data there.
	# If not, we add the data 1 index after the previous place the data was added to.
	else:
		index = 0
		for j in range(len(fs_list)):
			for i in range(len(all_fs_info)):
				next = False
				if all_fs_info[i]["Tag"] == fs_list[j]["Tag"] or all_fs_info[i]["Text"] == fs_list[j]["Text"]:
					all_fs_info[i][cur_year] = fs_list[j][cur_year]
					index = i+1
					next = True
					break
			if next: continue
			all_fs_info.insert(index, fs_list[j])
	
	return all_fs_info

def get_tags_from_df(fs_df):
	usgaap_tags = list(fs_df["Tag"])
	just_tags = list()
	for full_tag in usgaap_tags:
		tag = full_tag.split(':')[1]
		just_tags.append(tag)
	return just_tags

def process_fs(cfiles, fs, mag):
	all_fs_info = []
	print(f"⏳ Attempting to parse {fs} for {directory_cfiles_10K[i]}.")
	try:
		FS = fs_process_from_cfiles(cfiles, fs, False)
		fs_list = get_fs_list(FS, mag, fs)
		all_fs_info = populate_fs_df(fs_list, all_fs_info)
		print(f"	✅ Parsed successfully.\n")
	except:
		print(f"	❌ Could not parse {fs} for {directory_cfiles_10K[i]}.\n")
	
	return all_fs_info

def get_epv_info_from_fs(epv_info, epv_tags, fs_list):
	for tag_info in fs_list:
		tag = tag_info['Tag'].split(":")[1]
		for key in epv_tags:
			epv_tag = epv_tags[key]
			if tag in epv_tag:
				if tag in epv_info[cur_year]:
					epv_info[cur_year][key] += tag_info[cur_year]
				else:
					epv_info[cur_year][key] = tag_info[cur_year]

if __name__ == "__main__":
	ticker = sys.argv[1]
	mag = sys.argv[2]
	industry = sys.argv[3]
	
	div = 1000
	if mag == 't':
		div = 1000
	elif mag == 'm':
		div = 1000000
	elif div == 'n':
		div = 1

	offline = False
	if len(sys.argv) > 4:
		offline = True
	
	if not offline:
		CIK = get_company_CIK(ticker)
		# Retrieve the forms
		all_inline_10k_forms = get_forms_of_type_xbrl(CIK,'10-K', True)
		all_inline_10q_forms = get_forms_of_type_xbrl(CIK,'10-Q', True)
		
		if all_inline_10k_forms != []: retrieved_forms = save_all_forms(ticker,'10-K',all_inline_10k_forms)
		if all_inline_10q_forms != []: retrieved_forms = save_all_forms(ticker,'10-Q',all_inline_10q_forms)

	# Get the directory where the forms are/were stored and sort them in chronological order.1
	directory_cfiles_10Q = sorted(find_all_form_dir(ticker,"10-Q"), reverse=True)
	directory_cfiles_10K = sorted(find_all_form_dir(ticker,"10-K"), reverse=True)
	
	if directory_cfiles_10K[0] > directory_cfiles_10Q[0]:
		cfiles = read_forms_from_dir(f"forms/{ticker}/10-K/{directory_cfiles_10K[0]}")
	else:
		cfiles = read_forms_from_dir(f"forms/{ticker}/10-Q/{directory_cfiles_10Q[0]}")
	
	# Get the Balance Sheet information from the most recent 10-K/10-Q.
	NAV = fs_process_from_cfiles(cfiles, 'bs', True)
	balance_sheet = df_input_fs(NAV, div)
	df_nav = pd.DataFrame(balance_sheet)

	#TODO: I need to add retrieve both current and previous year bs info. Rn, im only getting the current year.
	# Attempt to parse as much fs info and append them all into all_fs_info
	all_bs_info = []
	all_is_info = []
	all_cf_info = []
	epv_info = {}
	epv_tags = json.loads(read_file(f"./tags/epv_{industry}_tags.json"))

	directory_cfiles_10K.reverse()
	for i in range(len(directory_cfiles_10K)):
		cur_year = directory_cfiles_10K[i]
		if cur_year == ".DS_Store":continue
		epv_info[cur_year] = {}
		cfiles = read_forms_from_dir(f"forms/{ticker}/10-K/{directory_cfiles_10K[i]}")
		
		print(f"⏳ Attempting to parse balance sheet for {directory_cfiles_10K[i]}.")
		try:
			# For getting all balance sheet info
			BS = fs_process_from_cfiles(cfiles, 'bs', False)
			bs_list = get_fs_list(BS, mag, 'bs')
			all_bs_info = populate_fs_df(bs_list, all_bs_info)

			# For getting the EPV info
			get_epv_info_from_fs(epv_info, epv_tags, bs_list)

			print(f"	✅ Parsed successfully.\n")
		except:
			print(f"	❌ Could not parse balance sheet for {directory_cfiles_10K[i]}.\n")

		print(f"⏳ Attempting to parse income statement for {directory_cfiles_10K[i]}.")
		try:
			IS = fs_process_from_cfiles(cfiles, 'is', False)
			is_list = get_fs_list(IS, mag, 'is')
			all_is_info = populate_fs_df(is_list, all_is_info)

			# For getting the EPV info
			get_epv_info_from_fs(epv_info, epv_tags, is_list)

			print(f"	✅ Parsed successfully.\n")
		except:
			print(f"	❌ Could not parse income statement for {directory_cfiles_10K[i]}.\n")

		print(f"⏳ Attempting to parse cash flow for {directory_cfiles_10K[i]}.")	
		try:
			CF = fs_process_from_cfiles(cfiles, 'cf', False)
			cf_list = get_fs_list(CF, mag, 'cf')
			all_cf_info = populate_fs_df(cf_list, all_cf_info)

			# For getting the EPV info
			get_epv_info_from_fs(epv_info, epv_tags, cf_list)
			
			print(f"	✅ Parsed successfully.\n")
		except:
			print(f"	❌ Could not parse cash flow for {directory_cfiles_10K[i]}.\n")

	df_bs = pd.DataFrame(all_bs_info)
	df_is = pd.DataFrame(all_is_info)
	df_cf = pd.DataFrame(all_cf_info)
			
	appeared_tags = list()
	try:
		appeared_tags = get_tags_from_df(df_bs) + get_tags_from_df(df_is) + get_tags_from_df(df_cf)
	except:
		appeared_tags  = []
	#TODO: Work on duplicate tags.
	
	# Retrieve facts json
	if not offline:
		retreieved_facts = save_all_facts(sys.argv[1])
	else:
		retreieved_facts = True
	if retreieved_facts:
		TAGS = get_tags(f"./forms/{ticker}/{ticker}.json", "10-K")
		EPV = epv(f"./tags/epv_{industry}_tags.json", f"./forms/{ticker}/{ticker}.json", appeared_tags)
	else:
		print("I have to work on facts from BS. Facts retreival failed as well. look into why.")

	writer = pd.ExcelWriter(f"./excel/{ticker}.xlsx", engine='xlsxwriter')
	df_nav.to_excel(writer, sheet_name='NAV')
	df_bs.to_excel(writer, sheet_name='Balance Sheet')
	df_is.to_excel(writer, sheet_name='Income Statement')
	df_cf.to_excel(writer, sheet_name='Cash Flow')
	TAGS.to_excel(writer, sheet_name='Tags')
	EPV.to_excel(writer, sheet_name='EPV Data')
	writer.save()
	#writer.close()

	path = f"./excel/{ticker}.xlsx"
	wb = xl.load_workbook(path)
	source = wb['Balance Sheet']
	target = wb.copy_worksheet(source)
	target.title = "NAV BS"

	col = 5
	max_col = target.max_column
	while col <= max_col:
		target.insert_cols(col)
		target.insert_cols(col)
		col += 3
		max_col += 2

	wb.create_sheet("COVER")
	wb_cover = wb["COVER"]
	fill_cover(wb_cover, ticker, 5)

	wb.create_sheet("WACC")
	wb_WACC = wb["WACC"]
	wacc_titles(wb_WACC)

	wb.create_sheet("EPV")
	wb_EPV = wb["EPV"]
	fill_epv(wb_EPV, epv_info)

	years = [date.split('-')[0] for date in epv_info]
	wb.create_sheet("GV")
	wb_GV = wb["GV"]
	fill_gv(wb_GV, years)

	wb.save(path)




	#TODO: EPS is not working because it thinks it is the same as the shares outstanding since the text is the same. Make if it is EPS,
	# we don't override.