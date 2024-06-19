import os
from bs4 import BeautifulSoup
import sys
from dataclasses import dataclass
from typing import List
import pandas as pd
import json
import openpyxl as xl

from files import read_forms_from_dir, find_latest_form_dir, find_all_form_dir
from xbrl_parse import get_fs_fields, get_disclosure_fields, get_diluted_common_shares_outstanding
from edgar_retrieve import get_company_CIK, get_forms_of_type_xbrl, save_all_facts, save_all_forms
from html_parse import html_to_facts
from html_process import derived_fs_table, assign_HTMLFact_to_XBRLNode
from util.write_to_csv import *
from util.file_management import read_file
from excel_model.epv import *
from excel_model.cover import *
from excel_model.gv import *
from excel_model.wacc import *
from excel_model.nav import *

def get_fs_list(fs_fields, mag, fs, date):
	div = 1000
	if mag == 't':
		div = 1000
	elif mag == 'm':
		div = 1000000
	elif div == 'n':
		div = 1

	fs_tag = ""
	if fs == 'bs': fs_tag = "us-gaap:Assets"

	elif fs == 'is': 
		if "us-gaap:NetIncomeLoss" in fs_fields:
			fs_tag = "us-gaap:NetIncomeLoss"
		elif "us-gaap:ProfitLoss" in fs_fields:
			fs_tag = "us-gaap:ProfitLoss"
		elif "us-gaap:OperatingIncomeLoss" in fs_fields:
			fs_tag = "us-gaap:OperatingIncomeLoss"

	elif fs == 'cf':
		if "us-gaap:NetIncomeLoss" in fs_fields:
			fs_tag = "us-gaap:NetIncomeLoss"
		elif "us-gaap:ProfitLoss" in fs_fields:
			fs_tag = "us-gaap:ProfitLoss"
		elif "us-gaap:OperatingIncomeLoss" in fs_fields:
			fs_tag = "us-gaap:OperatingIncomeLoss"
		elif "us-gaap:NetCashProvidedByUsedInOperatingActivitiesContinuingOperations" in fs_fields:
			fs_tag = "us-gaap:NetCashProvidedByUsedInOperatingActivitiesContinuingOperations"
		elif "us-gaap:NetCashProvidedByUsedInOperatingActivities" in fs_fields:
			fs_tag = "us-gaap:NetCashProvidedByUsedInOperatingActivities"
	
	if fs_tag == "": print(f"{fs} beginning tag is not found.")
	#assert fs_tag != "", "The base tag is not found in the financial statement. Look for another tag that'll be in it."
	#date = fs_fields[fs_tag].date[0]

	fs_list = []
	for key in fs_fields:
		if fs_fields[key].val is None:
			continue
		if date not in fs_fields[key].date:
			continue
		i_index = fs_fields[key].date.index(date)
		text = fs_fields[key].text[0] if fs_fields[key].text else ''
		if len(fs_fields[key].val) > i_index:
			d = {'tag': fs_fields[key].tag, 'Text':text, date: fs_fields[key].val[i_index]/div}
			fs_list.append(d)

		'''
		for i,tag in enumerate(fs_fields[key].text):
			if fs_fields[key].tag == "us-gaap:EarningsPerShareBasic" or fs_fields[key].tag == "us-gaap:EarningsPerShareDiluted":
				d = {'tag': fs_fields[key].tag, 'Text':text, date: fs_fields[key].val[i]}	
			else:
				d = {'tag': fs_fields[key].tag, 'Text':text, date: fs_fields[key].val[i]/div}
			fs_list.append(d)
		'''
	return fs_list

def get_all_dates(data, period, currency):
	dates1 = []
	cash = 'CashAndCashEquivalentsAtCarryingValue'
	if cash in data: 
		if currency in data[cash]['units']:
			d = list(data[cash]['units'][currency])
			for year in d:
				if year['form'] == period:
					if year['end'] not in dates1: dates1.append(year['end'])
	dates1.sort()

	dates2 = []
	cash = 'CashAndDueFromBanks'
	if cash in data: 
		if currency in data[cash]['units']:
			d = list(data[cash]['units'][currency])
			for year in d:
				if year['form'] == period:
					if year['end'] not in dates1: dates1.append(year['end'])

	dates3 = []
	cash = 'Cash'
	if cash in data: 
		if currency in data[cash]['units']:
			d = list(data[cash]['units'][currency])
			for year in d:
				if year['form'] == period:
					if year['end'] not in dates3: dates3.append(year['end'])

	dates1.sort()
	dates2.sort()
	dates3.sort()
	dates = dates1 if len(dates1) > len(dates2) else dates2
	dates = dates if len(dates) > len(dates3) else dates3
	return dates
	
def get_tags(destination, period):
	div = 1000
	if mag == 't':
		div = 1000
	elif mag == 'm':
		div = 1000000
	elif div == 'n':
		div = 1

	with open(destination, 'r') as f:
		data = f.read()
	facts_json = json.loads(data)['facts']['us-gaap']
	with open("./tags/facts_tags.json", 'r') as f:
		data = f.read()
	USGAAP_json = json.loads(data)

	dates = get_all_dates(facts_json, period,'USD')
	# Special cases where companies like CP has 'CAD' as currecny
	if (dates == []):
		dates = get_all_dates(facts_json, period,'CAD')

	df = pd.DataFrame(columns = dates)

	for category in USGAAP_json:
		for usgaap_tag in USGAAP_json[category]:
			if usgaap_tag in facts_json: 
				rowData = {y:0 for y in dates}
				if 'USD' in facts_json[usgaap_tag]['units']:
					d = list(facts_json[usgaap_tag]['units']['USD'])
					for year in d:
						if year['form'] == period and year['end'] in dates:
							if category == "EPS" or category == "Dividend":
								rowData[year['end']] = year['val']
							else:
								rowData[year['end']] = year['val']/div
				appendRow = [rowData[key] for key in rowData]
				df.loc[usgaap_tag] = appendRow
	return df

def epv(epv_path, json_path):
	with open(json_path, 'r') as f:
		data = f.read()
	data = json.loads(data)['facts']['us-gaap']
	usd = (list(data)[0])
	if 'USD' not in data[usd]['units']:
		return pd.DataFrame()
	
	with open(epv_path, 'r') as f:
		all_tags = f.read()
	all_tags = json.loads(all_tags)

	dates = get_all_dates(data, '10-K', 'USD')

	df = pd.DataFrame(columns = dates)
	for tags in all_tags:
		all_usgaap_rowData = []
		rowData = {y:0 for y in dates}
		for tag in all_tags[tags]:
			#if tag in data and tag in appeared_tags: 
			if tag in data: 
				if 'USD' in data[tag]['units']:
					d = list(data[tag]['units']['USD'])
					for year in d:
						if year['form'] == '10-K' and year['end'] in dates:
							#TODO: I need to handle fields that adds up to one category. For example debt usually has multiple that adds up to one.
							if rowData[year['end']] != 0: 
								continue 
							rowData[year['end']] += year['val']/div

							# Adding this part to also export all of the elements that sums up to the calulation parameter for EPV
							if {tag: rowData} not in all_usgaap_rowData:
								all_usgaap_rowData.append({tag: rowData})

		# Append the new category in EPV
		appendRow = [rowData[key] for key in rowData]			
		df.loc[tags] = appendRow

		# Also show all the us-gaap tags that calculates up to it
		for usgaap_rowData in all_usgaap_rowData:
			for usgaap_tag in usgaap_rowData:
				rowData = usgaap_rowData[usgaap_tag]
				appendRow = [rowData[key] for key in rowData]			
				df.loc[usgaap_tag] = appendRow
		
	#df.to_csv("ticker.csv")
	return df
					
def fs_process_from_cfiles(cfiles, fs, iGetIndexUpTo = 0, bIsChronological = [False]):
	# Get the Balance Sheet information from the most recent 10-K/10-Q.
	fs_fields = get_fs_fields(ticker, fs, cfiles)
	all_tables = html_to_facts(cfiles.html, cfiles.htm_xml, fs_fields)
	fs_table_from_html = derived_fs_table(all_tables, fs_fields)
	iNumOfFSColumns = len(fs_table_from_html[0])
	assert iNumOfFSColumns > 0, "Facts retrieved for the Financial Statement is empty."

	# If there is only one column, just get that and return
	if iNumOfFSColumns == 1:
		assign_HTMLFact_to_XBRLNode(fs_fields, fs_table_from_html, 0)
		return fs_fields
	
	# If greater, we want to identify if it is chronological or not
	date0 = fs_table_from_html[0][0].date
	date1 = fs_table_from_html[0][1].date
	bIsChronological[0] = False if date0 > date1 else True	
	if bIsChronological[0]:
		for i in range(0, iGetIndexUpTo+1):
			if i >= iNumOfFSColumns: break
			assign_HTMLFact_to_XBRLNode(fs_fields, fs_table_from_html, i)
	else:
		fs_fields = get_fs_fields(ticker, fs, cfiles)
		for i in range(iGetIndexUpTo, -1, -1):
			if i >= iNumOfFSColumns: continue
			assign_HTMLFact_to_XBRLNode(fs_fields, fs_table_from_html, i)
	return fs_fields

def df_input_fs(FS, dates, div=1000):
	financial_statement = []
	for date in dates:
		for tag in FS:
			tag_info = FS[tag]
			if tag_info.val is None or tag_info.date is None: continue
			indexes = [i for i, d in enumerate(tag_info.date) if d == date]
			val_date_dict = {tag_info.date[i]: tag_info.val[i]/div for i in indexes}
			#val_date_dict = {tag_info.date[i]: tag_info.val[i]/div for i in range(len(tag_info.date))}
			push_fact = {"tag": tag_info.text[0]}
			for key in val_date_dict:
				push_fact[key] = val_date_dict[key]

			financial_statement.append(push_fact)
	return financial_statement

def populate_fs_df(fs_list, all_fs_info, cur_year):
	if cur_year is None:
		cur_year = list(fs_list[0].keys())[-1]
	if all_fs_info == []: 
		all_fs_info = fs_list
	# Loop through the new list and add it to the right place.
	# First, we check of the tag is already added. If so, we just add the data there.
	# If not, we add the data 1 index after the previous place the data was added to.
	else:
		index = 0
		for j in range(len(fs_list)):
			for i in range(len(all_fs_info)):
				next = False
				if all_fs_info[i]["tag"] == fs_list[j]["tag"] or all_fs_info[i]["Text"] == fs_list[j]["Text"]:
					all_fs_info[i][cur_year] = fs_list[j][cur_year]
					index = i+1
					next = True
					break
			if next: continue
			all_fs_info.insert(index, fs_list[j])
	return all_fs_info

def get_tags_from_df(fs_df):
	usgaap_tags = list(fs_df["tag"])
	just_tags = list()
	for full_tag in usgaap_tags:
		tag = full_tag.split(':')[1]
		just_tags.append(tag)
	return just_tags

def get_epv_info_from_fs(epv_info, epv_tags, fs_list, cur_year):
	for tag_info in fs_list:
		tag = tag_info['tag'].split(":")[1]
		for key in epv_tags:
			epv_tag = epv_tags[key]
			if tag in epv_tag:
				if tag in epv_info[cur_year]:
					epv_info[cur_year][key] += tag_info[cur_year]
				else:
					if cur_year not in tag_info:
						continue
					epv_info[cur_year][key] = tag_info[cur_year]

def get_oldest_prior_year_revenue_from_is(all_is_info, first_10k_year, revenue_tags):
	oldest_prior_year_revene = 0
	for tag_info in all_is_info:
		if tag_info["tag"].split(':')[1] in revenue_tags:
			for key in tag_info:
				if key == "tag" or key == "Text":
					continue
				if key == first_10k_year:
					return oldest_prior_year_revene

				oldest_prior_year_revene = tag_info[key]
	return 0

def get_tag_info(all_fs_info, tag):
	for tag_info in all_fs_info:
		if tag == tag_info['tag']:
			return tag_info
	return None

def get_all_dates_from_fs_fields(FS, bIsChronological = True):
	dates = []
	for tag in FS:
		tag_info = FS[tag]
		cur_dates = tag_info.date
		if cur_dates is None: continue
		for date in cur_dates:
			if date not in dates: dates.append(date)
	dates.sort()
	if not bIsChronological:
		dates.reverse()
	return dates

def download_forms(ticker, offline=False):
	# Don't need to download if online
	if offline: 
		return
	
	CIK = get_company_CIK(ticker)
	# Retrieve the forms
	all_inline_10k_forms = get_forms_of_type_xbrl(CIK,'10-K', True)
	all_inline_10q_forms = get_forms_of_type_xbrl(CIK,'10-Q', True)
	
	if all_inline_10k_forms != []: save_all_forms(ticker,'10-K',all_inline_10k_forms)
	if all_inline_10q_forms != []: save_all_forms(ticker,'10-Q',all_inline_10q_forms)

def get_parsing_directories(ticker, parsing_method):
	directory_cfiles = []
	directory_cfiles_10Q = sorted(find_all_form_dir(ticker,"10-Q"))
	directory_cfiles_10Q = list(filter((".DS_Store").__ne__, directory_cfiles_10Q))
	directory_cfiles_10K = sorted(find_all_form_dir(ticker,"10-K"))
	directory_cfiles_10K = list(filter((".DS_Store").__ne__, directory_cfiles_10K))

	# Parse every single quarter and 10-K
	if parsing_method == 'q':
		directory_cfiles = sorted(directory_cfiles_10Q + directory_cfiles_10K)

	# Parse all 10-Q past the most recent 10-K
	elif parsing_method == 'r':
		directory_cfiles = directory_cfiles_10K + []	# Just adding + [] so it will not be a reference to directory_cfiles_10K
		for i, cfile_10Q in enumerate(directory_cfiles_10Q):
			if cfile_10Q > directory_cfiles_10K[-1]:
				directory_cfiles += directory_cfiles_10Q[i:]
				break
	# Parse the latest only
	# TODO: this will work once epv_info shit is sorted
	elif parsing_method == 'l':
		if directory_cfiles_10K[-1] > directory_cfiles_10Q[-1]:
			directory_cfiles = directory_cfiles_10K[-1:]
		else:
			directory_cfiles = directory_cfiles_10Q[-1:]
	# Only parse 10-K
	else:
		directory_cfiles = directory_cfiles_10K

	return directory_cfiles, directory_cfiles_10K, directory_cfiles_10Q

def run_main(ticker, mag, industry, parsing_method, risk_spread, beta, debt, ppe):
	if parsing_method == 'q':
		print("Parsing all 10-K and 10-Q...")
	
	div = 1000
	if mag == 't':
		div = 1000
	elif mag == 'm':
		div = 1000000
	elif div == 'n':
		div = 1

	offline = False
	if len(sys.argv) > 5:
		offline = True
	
	download_forms(ticker, offline)
	directory_cfiles, directory_cfiles_10K, directory_cfiles_10Q = get_parsing_directories(ticker, parsing_method)
	
	if directory_cfiles_10K[-1] > directory_cfiles_10Q[-1]:
		cfiles = read_forms_from_dir(f"forms/{ticker}/10-K/{directory_cfiles_10K[-1]}")
	else:
		cfiles = read_forms_from_dir(f"forms/{ticker}/10-Q/{directory_cfiles_10Q[-1]}")
	
	# Get the Balance Sheet information from the most recent 10-K/10-Q.
	try:
		NAV = fs_process_from_cfiles(cfiles, 'bs', 1)
		dates = get_all_dates_from_fs_fields(NAV)
		balance_sheet = df_input_fs(NAV, dates, div)
		df_nav = pd.DataFrame(balance_sheet)
	except:
		df_nav = {'None': {"Failed"}}
		print("Failed to process NAV.")

	#TODO: I need to add retrieve both current and previous year bs info. Rn, im only getting the current year.
	# Attempt to parse as much fs info and append them all into all_fs_info
	all_bs_info = []
	all_is_info = []
	all_cf_info = []
	epv_info = {}
	epv_tags = json.loads(read_file(f"./tags/epv_{industry}_tags.json"))
	shares_outsanding = {}
	
	#directory_cfiles = list(filter((".DS_Store").__ne__, directory_cfiles))
	for i, cur_year in enumerate(directory_cfiles):
		if cur_year in directory_cfiles_10K:
			epv_info[cur_year] = {}
		try:
			if cur_year in directory_cfiles_10K:
				cfiles = read_forms_from_dir(f"forms/{ticker}/10-K/{cur_year}")
			else:
				cfiles = read_forms_from_dir(f"forms/{ticker}/10-Q/{cur_year}")
		except:
			print(f"Missing file for {cur_year}.")

		# Parsing Balance Sheet
		print(f"⏳ Attempting to parse balance sheet for {cur_year}.")
		try:
			# For getting all balance sheet info
			BS = fs_process_from_cfiles(cfiles, 'bs', 0)
			bs_list = get_fs_list(BS, mag, 'bs', cur_year)
			all_bs_info = populate_fs_df(bs_list, all_bs_info, cur_year)

			# For getting the EPV info
			if cur_year in directory_cfiles_10K:
				get_epv_info_from_fs(epv_info, epv_tags, all_bs_info, cur_year)

			# Get Shares Outstanding
			str_shares_outstanding = get_diluted_common_shares_outstanding(cfiles.htm_xml)
			#TODO: gotta split because apprently you can't do int(str) if str has decimal. Look into see if there are more efficient ways.
			cur_shares_outstanding = str_shares_outstanding.split('.')[0]
			cur_shares_outstanding = int(cur_shares_outstanding)/div
			shares_outsanding[cur_year] = cur_shares_outstanding

			print(f"	✅ Parsed successfully.\n")
		except:
			print(f"	❌ Could not parse balance sheet for {cur_year}.\n")

		# Parsing Income Statement
		print(f"⏳ Attempting to parse income statement for {cur_year}.")
		# For now, we will only retrieve info from 10-K for IS
		if cur_year in directory_cfiles_10Q:
			print(f"	🔒 Skipping parsing for 10-Q for now.\n")
			continue
		try:
			if i == 0 or cur_year == directory_cfiles_10K[0]:
				index = 2
			else: 
				index = 0
			bIsChronological = [True]
			IS = fs_process_from_cfiles(cfiles, 'is', index, bIsChronological)
			dates = get_all_dates_from_fs_fields(IS, bIsChronological[0])
			for j,date in enumerate(dates):
				is_list = get_fs_list(IS, mag, 'is', date)
				all_is_info = populate_fs_df(is_list, all_is_info, date)
				#if bIsChronological[0] and j == len(IS)-1:
				#	all_is_info = populate_fs_df(is_list, all_is_info, cur_year)
				#elif not bIsChronological[0] and j == 0:
				#	all_is_info = populate_fs_df(is_list, all_is_info, cur_year)
				#else:
				#	all_is_info = populate_fs_df(is_list, all_is_info, None)

			# For getting the EPV info
			if cur_year in directory_cfiles_10K:
				get_epv_info_from_fs(epv_info, epv_tags, all_is_info, cur_year)

			print(f"	✅ Parsed successfully.\n")
		except:
			print(f"	❌ Could not parse income statement for {cur_year}.\n")

		# Parsing Cash Flow
		print(f"⏳ Attempting to parse cash flow for {cur_year}.")
		# For now, we will only retrieve info from 10-K for IS
		if cur_year in directory_cfiles_10Q:
			print(f"	 🔒 Skipping parsing for 10-Q for now.\n")
			continue
		try:
			CF = fs_process_from_cfiles(cfiles, 'cf', 0)
			cf_list = get_fs_list(CF, mag, 'cf', cur_year)
			all_cf_info = populate_fs_df(cf_list, all_cf_info, cur_year)

			# For getting the EPV info
			if cur_year in directory_cfiles_10K:
				get_epv_info_from_fs(epv_info, epv_tags, all_cf_info, cur_year)
			
			print(f"	✅ Parsed successfully.\n")
		except:
			print(f"	❌ Could not parse cash flow for {cur_year}.\n")

	iAsset, iLiabilities = None, None
	for i, info in enumerate(all_bs_info):
		if info["tag"] == "us-gaap:Assets":
			iAsset = i
			continue
		if info["tag"] == "us-gaap:Liabilities":
			iLiabilities = i
			continue
		if iLiabilities is not None and info["tag"] == "us-gaap:PreferredStockValue" or info["tag"] == "us-gaap:CommitmentsAndContingencies" or info["tag"] == "us-gaap:CommonStockValue":
			iLiabilities = i-1
			continue
	if iAsset is None:
		print("iAsset is not found")
	if iLiabilities is None:
		print("iLiabilities is not found")

	assets_info = all_bs_info[:iAsset+1]
	liabilities_info = all_bs_info[iAsset+1:iLiabilities+1]

	df_bs = pd.DataFrame(all_bs_info)
	df_is = pd.DataFrame(all_is_info)
	df_cf = pd.DataFrame(all_cf_info)

	writer = pd.ExcelWriter(f"./excel/{ticker}.xlsx", engine='xlsxwriter')
	df_bs.to_excel(writer, sheet_name='Balance Sheet')
	df_is.to_excel(writer, sheet_name='Income Statement')
	df_cf.to_excel(writer, sheet_name='Cash Flow')
	writer.save()

	years = [date.split('-')[0] for date in epv_info]
	dates = [date for date in epv_info]

	# Load in the SG&A tags
	SGA_tags = []
	with open('./tags/facts_tags.json') as f:
		file_contents = f.read()
		SGA_tags = json.loads(file_contents)["SG&A"]
	# Get the SG&A values
	SGA_values = {}
	for tag_info in all_is_info:
		if tag_info["tag"].split(':')[1] in SGA_tags:
			for key in tag_info:
				if key == "tag" or key == "Text" or key in directory_cfiles_10Q:
					continue
				if key in SGA_values:
					SGA_values[key] += tag_info[key]
				else:
					SGA_values[key] = tag_info[key]
	for date in dates:
		if date not in SGA_values:
			SGA_values[date] = 0

	myKeys = list(SGA_values.keys())
	myKeys.sort()
	SGA_values = {i: abs(SGA_values[i]) for i in myKeys}
	SGA_list = list(SGA_values.values())

	# Get all dates in Income Statement (for now)
	all_dates = []
	for tag_info in all_is_info:
		for key in tag_info:
			if key == 'tag' or key == 'Text': continue
			if key not in all_dates:
				all_dates.append(key)
	all_dates.sort()

	if parsing_method != 'l':
		first_epv_date = list(epv_info.keys())[0]
		i_first_epv_date = all_dates.index(first_epv_date)
		first_prior_year_revenue_date = all_dates[i_first_epv_date - 1]

		# Load the Revenue tags
		revenue_tags = []
		with open(f"./tags/epv_{industry}_tags.json") as f:
			file_contents = f.read()
			revenue_tags = json.loads(file_contents)["Current Year Revenue"]

		oldest_prior_year_revenue = 0
		for tag_info in all_is_info:
			for tag in revenue_tags:
				if tag in tag_info['tag']:
					if first_prior_year_revenue_date in tag_info:
						oldest_prior_year_revenue = tag_info[first_prior_year_revenue_date]
						break

	wb = xl.Workbook()
	wb_cover = wb["Sheet"]
	wb_cover.title = "COVER"
	fill_cover(wb_cover, ticker, 5)

	wb.create_sheet("WACC")
	wb_WACC = wb["WACC"]
	fill_wacc(wb_WACC, mag)
	wb.create_sheet("NAV")
	wb_NAV = wb["NAV"]

	NAV_summary_row, iNAVPriceCoord, i_quarter = fill_NAV(wb_NAV, assets_info, liabilities_info, shares_outsanding, directory_cfiles, directory_cfiles_10K, SGA_list, ppe)

	wb.create_sheet("EPV")
	wb_EPV = wb["EPV"]
	EPV_rows, iEPVPriceCoord = fill_epv(wb_EPV, industry, epv_info, NAV_summary_row.shares)

	iNAVRow = NAV_summary_row.shares-1
	wb.create_sheet("GV")
	wb_GV = wb["GV"]
	iGVPriceCoord = fill_gv(wb_GV, years, iNAVRow, EPV_rows)

	# Some hard coded stuff
	wb_cover.cell(row=2, column=2, value=ticker)
	wb_cover.cell(row=9, column=3, value=f"=NAV!{iNAVPriceCoord[0]+str(iNAVPriceCoord[1])}")
	wb_cover.cell(row=10, column=3, value=f"=EPV!{iEPVPriceCoord[0]+str(iEPVPriceCoord[1])}")
	wb_cover.cell(row=11, column=3, value=f"=GV!{iGVPriceCoord[0]+str(iGVPriceCoord[1])}")

	wb_WACC.cell(row=7, column=3, value=risk_spread)
	wb_WACC.cell(row=27, column=3, value=beta)

	if parsing_method != 'l':
		wb_WACC.cell(row=6, column=7, value=int(years[-1]))
		wb_EPV.cell(row=EPV_rows.get["Prior Year Revenue"], column=3).value = oldest_prior_year_revenue
		
		try:
			interest_tag_info = get_tag_info(all_is_info, 'us-gaap:InterestExpense')
			if interest_tag_info is None:
				interest_tag_info = get_tag_info(all_is_info, 'us-gaap:InterestAndDebtExpense')
			elif interest_tag_info is None:
				interest_tag_info = get_tag_info(all_is_info,'us-gaap:InterestIncomeExpenseNonoperatingNet')
			interest_expense = abs(interest_tag_info[directory_cfiles_10K[-1]])
			wb_WACC.cell(row=6, column=9, value=interest_expense)
		except:
			print("tag for interst expense is wrong")

	sBasis = "FY"
	if parsing_method == 'q':
		sBasis = "Q"

	year = int(directory_cfiles_10K[-1].split('-')[0])
	i_quarter -= 1
	if i_quarter > 0 and i_quarter < 4:
		year += 1
		sBasis = "Q"
	else:
		sBasis = "FY"
		i_quarter = ''
	#return wb
	wb.save(f"/Users/caseymackinnon/Desktop/Personalized Finance/edgar/excel/{ticker} - {sBasis}{i_quarter} {year} - {directory_cfiles[-1]}.xlsx")


	#TODO: EPS is not working because it thinks it is the same as the shares outstanding since the text is the same. Make if it is EPS,
	# we don't override.

if __name__ == "__main__":
	ticker = input("Enter Ticker: ")
	mag = input("Enter Reported Scale: ")
	industry = input("Enter Industry: ")
	print()
	risk_spread = input("Enter Risk Spread: ")
	beta = input("Enter Beta: ")
	debt = input("Enter Debt: ")
	ppe = input("Enter PPE Adjustment: ")

	parsing_method = 'r'
	run_main(ticker, mag, industry, parsing_method, risk_spread, beta, debt, ppe)
