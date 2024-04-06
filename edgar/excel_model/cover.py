import openpyxl as xl
from openpyxl import formatting
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.styles.numbers as format

CUSTOM_FORMAT_CURRENCY_ZERO = '_($* #,##0_);[Red]_($* (#,##0);_($* "-"??_)'
CUSTOM_FORMAT_CURRENCY_ONE = '_($* #,##0.0_);[Red]_($* (#,##0.0);_($* "-"??_)'
CUSTOM_FORMAT_CURRENCY_TWO = '_($* #,##0.00_);[Red]_($* (#,##0.00);_($* "-"??_)'
CUSTOM_FORMAT_PE = '0.00x'
letters = ' ABCDEFGHIJKLMNOPQRSTUVWXYZ'

boldFont = Font(name='Arial',size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
textFont  = Font(name='Arial',size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='000000')

orangeFill = PatternFill(fill_type="solid", start_color='f9a766', end_color='f9a766')
greenerFill = PatternFill(fill_type="solid", start_color='7ad593', end_color='7ad593')
greyFill = PatternFill(fill_type="solid", start_color='ebebeb', end_color='ebebeb')
yellowishFill = PatternFill(fill_type="solid", start_color='fef2cc', end_color='fef2cc')

thinBorder = Side(style="thin", color="000000")
thickBorder = Side(style="thick", color="000000")
noBorder = Side(style="none")

def fill_cover(wb_cover, ticker, iYears):
    wb_cover.column_dimensions['A'].width = 2
    # Titles
    row, col = 2, 2
    wb_cover.column_dimensions['B'].width = 40
    for _ in range(2, 7):
        wb_cover.cell(row=2, column=2, value=None)
        cell = wb_cover[f"{letters[col]}{row}"]
        cell.fill = orangeFill
        cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)

        if row == 2:
            cell.border = Border(left=thickBorder, top=thickBorder, right=noBorder, bottom=noBorder)
        if row == 3:
            wb_cover.cell(row=row, column=col, value="My WACC (discount rate)")
        if row == 4:
            wb_cover.cell(row=row, column=col, value="P/E")
        if row == 5:
            wb_cover.cell(row=row, column=col, value="Shares Outstanding (millions)")
        if row == 6:
            wb_cover.cell(row=row, column=col, value="Market Capitalization (in millions)")
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thickBorder)
        row += 1

    # Title Data
    row, col = 2, 3
    wb_cover.column_dimensions['C'].width = 15
    for _ in range(2, 7):
        wb_cover.cell(row=row, column=col, value=None)
        cell = wb_cover[f"{letters[col]}{row}"]
        cell.fill = greenerFill
        cell.font = boldFont
        cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=noBorder)

        if row == 2:
            wb_cover.cell(row=row, column=col, value=f"=B2.Price")
            cell.border = Border(left=thickBorder, top=thickBorder, right=thickBorder, bottom=noBorder)
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
        if row == 3:
            wb_cover.cell(row=row, column=col, value=f"=WACC!F20")
            cell.number_format = format.BUILTIN_FORMATS[10]
        if row == 4:
            wb_cover.cell(row=row, column=col, value=f"=B2.[P/E]")
            cell.number_format = CUSTOM_FORMAT_PE
        if row == 5:
            wb_cover.cell(row=row, column=col, value=f"=B2.[Shares outstanding]/1000000")
            cell.number_format = format.BUILTIN_FORMATS[2]
        if row == 6:
            wb_cover.cell(row=row, column=col, value=f"=B2.[Market cap]/1000000")
            cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=thickBorder)
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ZERO
        row += 1

    # Model Upside
    row, col = 8, 2
    for _ in range(8, 12):
        wb_cover.cell(row=row, column=col, value=None)
        cell = wb_cover[f"{letters[col]}{row}"]
        cell.font = boldFont
        cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)

        if row == 8:
            wb_cover.cell(row=row, column=col, value=None)
            cell.border = Border(left=thickBorder, top=thickBorder, right=noBorder, bottom=thinBorder)
        if row == 9:
            wb_cover.cell(row=row, column=col, value="NAV Upside")
        if row == 10:
            wb_cover.cell(row=row, column=col, value="EPV Upside")
        if row == 11:
            wb_cover.cell(row=row, column=col, value="GV Upside")
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thickBorder)
        row += 1
    
    # Price
    row, col = 8, 3
    for _ in range(8, 12):
        wb_cover.cell(row=row, column=col, value="Price")
        cell = wb_cover[f"{letters[col]}{row}"]
        cell.alignment = Alignment(horizontal="right")
        cell.font = boldFont
        cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)

        if row == 8:
            cell.font = boldFont
            cell.fill = greyFill
            cell.border = Border(left=thickBorder, top=thickBorder, right=noBorder, bottom=thinBorder)
        if row == 9:
            wb_cover.cell(row=row, column=col, value=f"=NAV!O60")
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
        if row == 10:
            wb_cover.cell(row=row, column=col, value=f"=EPV!{letters[4+iYears]}21")
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
        if row == 11:
            wb_cover.cell(row=row, column=col, value=f"=GV!{letters[4+iYears]}11")
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thickBorder)
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
        row += 1

    # Upside (%)
    wb_cover.column_dimensions['D'].width = 15
    row, col = 8, 4
    for _ in range(8, 12):
        wb_cover.cell(row=row, column=col, value="Upside (%)")
        cell = wb_cover[f"{letters[col]}{row}"]
        cell.alignment = Alignment(horizontal="right")
        cell.font = boldFont
        cell.border = Border(left=noBorder, top=noBorder, right=noBorder, bottom=noBorder)

        if row == 8:
            cell.font = boldFont
            cell.border = Border(left=noBorder, top=thickBorder, right=noBorder, bottom=thinBorder)
            cell.fill = greyFill
        if row == 9:
            wb_cover.cell(row=row, column=col, value="=C9/C2-1")
            cell.number_format = format.FORMAT_PERCENTAGE_00
        if row == 10:
            wb_cover.cell(row=row, column=col, value="=C10/C2-1")
            cell.number_format = format.FORMAT_PERCENTAGE_00
        if row == 11:
            wb_cover.cell(row=row, column=col, value="=C11/C2-1")
            cell.border = Border(left=noBorder, top=thinBorder, right=noBorder, bottom=thickBorder)
            cell.number_format = format.FORMAT_PERCENTAGE_00
        row += 1

    # Upside ($)
    wb_cover.column_dimensions['E'].width = 15
    row, col = 8, 5
    for _ in range(8, 12):
        wb_cover.cell(row=row, column=col, value="Upside ($)")
        cell = wb_cover[f"{letters[col]}{row}"]
        cell.alignment = Alignment(horizontal="right")
        cell.font = boldFont
        cell.border = Border(left=noBorder, top=noBorder, right=noBorder, bottom=noBorder)

        if row == 8:
            cell.font = boldFont
            cell.border = Border(left=noBorder, top=thickBorder, right=noBorder, bottom=thinBorder)
            cell.fill = greyFill
        if row == 9:
            wb_cover.cell(row=row, column=col, value="=C9-C2")
            cell.number_format = format.FORMAT_CURRENCY_USD_SIMPLE
        if row == 10:
            wb_cover.cell(row=row, column=col, value="=C10-C2")
            cell.number_format = format.FORMAT_CURRENCY_USD_SIMPLE
        if row == 11:
            wb_cover.cell(row=row, column=col, value="=C11-C2")
            cell.border = Border(left=noBorder, top=noBorder, right=noBorder, bottom=thickBorder)
            cell.number_format = format.FORMAT_CURRENCY_USD_SIMPLE
        row += 1

    # 30% Margin
    wb_cover.column_dimensions['F'].width = 20
    row, col = 8, 6
    for _ in range(8, 12):
        wb_cover.cell(row=row, column=col, value="30% Margin Price")
        cell = wb_cover[f"{letters[col]}{row}"]
        cell.alignment = Alignment(horizontal="right")
        cell.font = boldFont
        cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        cell.fill = yellowishFill

        if row == 8:
            cell.font = boldFont
            cell.border = Border(left=noBorder, top=thickBorder, right=thickBorder, bottom=thinBorder)
            cell.fill = greyFill
        if row == 9:
            wb_cover.cell(row=row, column=col, value="=C9/1.3")
            cell.number_format = format.FORMAT_CURRENCY_USD_SIMPLE
        if row == 10:
            wb_cover.cell(row=row, column=col, value="=C10/1.3")
            cell.number_format = format.FORMAT_CURRENCY_USD_SIMPLE
        if row == 11:
            wb_cover.cell(row=row, column=col, value="=C11/1.3")
            cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=thickBorder)
            cell.number_format = format.FORMAT_CURRENCY_USD_SIMPLE
        row += 1
    
    #ws.conditional_formatting.add('D9:D11', formatting.rule.CellIsRule(operator='lessThan', formula=['0'], fill=red_fill, font=red_font))