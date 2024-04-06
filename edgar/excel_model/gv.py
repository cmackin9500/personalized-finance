import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.styles.numbers as format

CUSTOM_FORMAT_CURRENCY_ONE = '_($* #,##0.0_);[Red]_($* (#,##0.0);_($* "-"??_)'
CUSTOM_FORMAT_CURRENCY_TWO = '_($* #,##0.00_);[Red]_($* (#,##0.00);_($* "-"??_)'
CUSTOM_FORMAT_PE = '0.00x'
letters = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 
           15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z', 27: 'AA', 
           28: 'AB', 29: 'AC', 30: 'AD', 31: 'AE', 32: 'AF', 33: 'AG', 34: 'AH', 35: 'AI', 36: 'AJ', 37: 'AK', 38: 'AL', 39: 'AM', 
           40: 'AN', 41: 'AO', 42: 'AP', 43: 'AQ', 44: 'AR', 45: 'AS', 46: 'AT', 47: 'AU', 48: 'AV', 49: 'AW', 50: 'AX', 51: 'AY', 
           52: 'AZ', 53: 'BA', 54: 'BB', 55: 'BC', 56: 'BD', 57: 'BE', 58: 'BF', 59: 'BG', 60: 'BH', 61: 'BI', 62: 'BJ', 63: 'BK', 
           64: 'BL', 65: 'BM', 66: 'BN', 67: 'BO', 68: 'BP', 69: 'BQ', 70: 'BR', 71: 'BS', 72: 'BT', 73: 'BU', 74: 'BV', 75: 'BW', 
           76: 'BX', 77: 'BY', 78: 'BZ'
           }
gv_titles = ["Earnings", "Net Asset Value", "Return on Net Asset Value", "Cost of Equity", "Growth Multiple (x)", "EPV", "Growth Value", 
             "Shares outstanding", "Growth Value Per Share", "Current Share Price", "Upside"]

boldFont = Font(name='Arial',size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
textFont  = Font(name='Arial',size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='000000')

orangeFill = PatternFill(fill_type="solid", start_color='fde599', end_color='fde599')
orangerFill = PatternFill(fill_type="solid", start_color='f6b26b', end_color='f6b26b')
greenerFill = PatternFill(fill_type="solid", start_color='b6d7a8', end_color='b6d7a8')
greenerFill = PatternFill(fill_type="solid", start_color='93c47d', end_color='93c47d')
greyFill = PatternFill(fill_type="solid", start_color='ebebeb', end_color='ebebeb')
greyerFill = PatternFill(fill_type="solid", start_color='999997', end_color='999997')
yellowFill = PatternFill(fill_type="solid", start_color='ffff01', end_color='ffff01')
notesFill = PatternFill(fill_type="solid", start_color='fdd966', end_color='fdd966')
titleFill = PatternFill(fill_type="solid", start_color='ebebeb', end_color='ebebeb')

thinBorder = Side(style="thin", color="000000")
thickBorder = Side(style="thick", color="000000")
noBorder = Side(style="none")

def GV_titles(wb_GV, row, col):
    wb_GV.column_dimensions['B'].width = 30
    wb_GV.cell(row=2, column=2, value="")
    cell = wb_GV["B2"]
    cell.fill = greyerFill
    cell.border = Border(left=thickBorder, top=thickBorder, right=noBorder, bottom=thinBorder)

    row = 3
    for i, value in enumerate(gv_titles):
        wb_GV.cell(row=row, column=col, value=value)
        cell = wb_GV[f"B{row}"]
        cell.font = boldFont
        cell.fill = greyFill
        
        if (row == 10):
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thinBorder)
        elif (row == 13):
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thickBorder)
        else:
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        row += 1

def fill_gv_data(wb_GV, row, col, years, iNAVRow):
    col = 3
    for i, year in enumerate(years):
        row = 2
        wb_GV.column_dimensions[letters[col]].width = 15
        for _ in range(2, 14):
            wb_GV.cell(row=row, column=col, value=None)
            cell = wb_GV[f"{letters[col]}{row}"]
            cell.font = textFont
            cell.fill = greenerFill
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            cell.border = Border(left=thinBorder, top=noBorder, right=thinBorder, bottom=noBorder)           

            # Year
            if row == 2:
                wb_GV.cell(row=row, column=col, value=year)
                cell.border = Border(left=thinBorder, top=thickBorder, right=thinBorder, bottom=thickBorder)
                cell.font = boldFont
                cell.alignment = Alignment(horizontal="center")
                cell.fill = greenerFill
                cell.number_format = format.FORMAT_NUMBER
            # Earnings
            if row == 3:
                wb_GV.cell(row=row, column=col, value=f"=EPV!{letters[col]}3")
            # Net Asset Value
            elif row  == 4:
                wb_GV.cell(row=row, column=col, value=f"=NAV!{letters[3+i*3]}{iNAVRow}")
            # Return on Net Asset Value
            elif row == 5:
                wb_GV.cell(row=row, column=col, value=f"={letters[col]}3/{letters[col]}4")
                cell.number_format = CUSTOM_FORMAT_PE
            # Cost of Equity
            elif row == 6:
                wb_GV.cell(row=row, column=col, value=0.12)
                cell.number_format = format.FORMAT_PERCENTAGE
            # Growth Multiple
            elif row == 7:
                wb_GV.cell(row=row, column=col, value=f"={letters[col]}5/{letters[col]}6")
                cell.number_format = CUSTOM_FORMAT_PE
             # EPV
            elif row == 8:
                wb_GV.cell(row=row, column=col, value=f"=EPV!{letters[col]}19")          
            # Growth Value
            elif row == 9:
                wb_GV.cell(row=row, column=col, value=f"={letters[col]}7*{letters[col]}8")
            # Shares Outstanding
            elif row == 10:
                wb_GV.cell(row=row, column=col, value=f"=EPV!{letters[col]}20")
                cell.border = Border(left=thinBorder, top=noBorder, right=thinBorder, bottom=thickBorder)
                cell.number_format = format.FORMAT_NUMBER_00
            # Growth Value Per Share
            elif row == 11:
                wb_GV.cell(row=row, column=col, value=f"={letters[col]}9/{letters[col]}10")
                cell.font = boldFont
                cell.fill = yellowFill
                cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
            # Current Share Price
            elif row == 12:
                wb_GV.cell(row=row, column=col, value="=COVER!C2")
                cell.font = boldFont
                cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
            # Upside
            elif row == 13:
                wb_GV.cell(row=row, column=col, value=f"={letters[col]}11/{letters[col]}12-1")
                cell.border = Border(left=thinBorder, top=noBorder, right=thinBorder, bottom=thickBorder)
                cell.number_format = format.FORMAT_PERCENTAGE
            row += 1
        col += 1

def GV_notes(wb_GV, row, col):
    wb_GV.column_dimensions[letters[col]].width = 15
    cell = wb_GV[f"{letters[col]}2"]
    cell.fill = greyFill
    cell.border = Border(left=thickBorder, top=thickBorder, right=noBorder, bottom=thinBorder)

    row = 2
    for _ in range(2, 14):
        wb_GV.cell(row=row, column=col, value=None)
        cell = wb_GV[f"{letters[col]}{row}"]
        cell.fill = orangeFill
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=thinBorder, top=noBorder, right=thickBorder, bottom=noBorder)

        if (row == 2):
            cell.fill = orangerFill
            cell.border = Border(left=thinBorder, top=thickBorder, right=thickBorder, bottom=thinBorder)
        elif (row == 3):
            wb_GV.cell(row=row, column=col, value="(a)")
        elif (row == 4):
            wb_GV.cell(row=row, column=col, value="(b)")
        elif (row == 5):
            wb_GV.cell(row=row, column=col, value="(c) = (a)/(b)")
        elif (row == 6):
            wb_GV.cell(row=row, column=col, value="(d)")
        elif (row == 7):
            wb_GV.cell(row=row, column=col, value="(e) = (c) / (d)")
        elif (row == 8):
            wb_GV.cell(row=row, column=col, value="(f)")
        elif (row == 9):
            wb_GV.cell(row=row, column=col, value="(g) = (e) x (f)")
        elif (row == 10):
            wb_GV.cell(row=row, column=col, value="(h)")
            cell.border = Border(left=thinBorder, top=noBorder, right=thickBorder, bottom=thinBorder)
        elif (row == 11):
            wb_GV.cell(row=row, column=col, value="(i) = (g) / (h)")
        elif row == 13:
            cell.border = Border(left=thinBorder, top=noBorder, right=thickBorder, bottom=thickBorder)
        row += 1

def fill_gv(wb_GV, years, iNAVRow):
    wb_GV.column_dimensions['A'].width = 2
    row, col = 2,2
    GV_titles(wb_GV, row, col)
    fill_gv_data(wb_GV, row, col, years, iNAVRow)
    col = 3 + len(years)
    GV_notes(wb_GV, row, col)