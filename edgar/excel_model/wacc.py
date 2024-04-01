import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.styles.numbers as format

CUSTOM_FORMAT_CURRENCY_ZERO = '_($* #,##0_);[Red]_($* (#,##0);_($* "-"??_)'
CUSTOM_FORMAT_CURRENCY_ONE = '_($* #,##0.0_);[Red]_($* (#,##0.0);_($* "-"??_)'
CUSTOM_FORMAT_CURRENCY_TWO = '_($* #,##0.00_);[Red]_($* (#,##0.00);_($* "-"??_)'
CUSTOM_FORMAT_PE = '0.00x'
RISK_SPREAD = '=SWITCH(B6, "Aaa", 0.63%,"AAA",0.63%,"AA",0.78%,"Aa2",0.78%,"A1",0.98%,"A+",0.98%,"A",1.08%,"A2",1.08%,"A3",1.22%,"A-",1.22%,"BBB",1.56%,"Baa2",1.56%,"Ba1",2%,"BB+",2%,"Ba2",2.4%,"BB",2.4%,"B1",3.51%,"B+",3.51%,"B2",4.21%,"B",4.21%,"B3",5.15%,"B-",5.15%,"Caa",8.2%,"CCC",8.2%,"Ca2",8.64%,"CC",8.64%,"C2",11.34%,"C",11.34%,"D2",15.12%,"D",15.12%,"AA+",0.71%,"Aa1",0.71%,"Aa3",0.88%,"AA-",0.88%,"BBB+",1.39%,"Baa1",1.39%,"BBB-",1.78%,"Baa3",1.78%,"BB-",2.96%,"Ba3",2.96%,2%)'
letters = ' ABCDEFGHIJKLMNOPQRSTUVWXYZ'
gv_titles = ["Earnings", "Net Asset Value", "Return on Net Asset Value", "Cost of Equity", "Growth Multiple (x)", "EPV", "Growth Value", 
             "Shares outstanding", "Growth Value Per Share", "Current Share Price", "Upside"]

boldFont = Font(name='Arial',size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
boldRedFont = Font(name='Arial',size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF0000')
textFont  = Font(name='Arial',size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='000000')

orangeFill = PatternFill(fill_type="solid", start_color='fde599', end_color='fde599')
orangerFill = PatternFill(fill_type="solid", start_color='f6b26b', end_color='f6b26b')
greenFill = PatternFill(fill_type="solid", start_color='b6d7a8', end_color='b6d7a8')
greenerFill = PatternFill(fill_type="solid", start_color='93c47d', end_color='93c47d')
greyFill = PatternFill(fill_type="solid", start_color='ebebeb', end_color='ebebeb')
greyerFill = PatternFill(fill_type="solid", start_color='999997', end_color='999997')
yellowFill = PatternFill(fill_type="solid", start_color='ffff01', end_color='ffff01')
notesFill = PatternFill(fill_type="solid", start_color='fdd966', end_color='fdd966')
titleFill = PatternFill(fill_type="solid", start_color='ebebeb', end_color='ebebeb')

thinBorder = Side(style="thin", color="000000")
thickBorder = Side(style="thick", color="000000")
noBorder = Side(style="none")

def create_box(wb_WACC, x, y, dx, dy):
    #create_box(wb_WACC, 5, 13, 5, 8)
    for col in range(x,x+dx):
        cell = wb_WACC[f"{letters[col]}{y}"]
        cell.border = Border(left=noBorder, top=thinBorder, right=noBorder, bottom=noBorder)
        cell = wb_WACC[f"{letters[col]}{y+dy-1}"]
        cell.border = Border(left=noBorder, top=noBorder, right=noBorder, bottom=thinBorder)
    for row in range(y,y+dy):
        cell_left = wb_WACC[f"{letters[x]}{row}"]
        cell_right = wb_WACC[f"{letters[x+dx-1]}{row}"]
        if row == y:
            cell_left.border = Border(left=thinBorder, top=thinBorder, right=noBorder, bottom=noBorder)
            cell_right.border = Border(left=noBorder, top=thinBorder, right=thinBorder, bottom=noBorder)
        elif row == y+dy-1:
            cell_left.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=thinBorder)
            cell_right.border = Border(left=noBorder, top=noBorder, right=thinBorder, bottom=thinBorder)
        else:
            cell_left.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
            cell_right.border = Border(left=noBorder, top=noBorder, right=thinBorder, bottom=noBorder)


def wacc_box_one(wb_WACC):
    create_box(wb_WACC, 2, 2, 8, 10)

    row = 2
    for col in range(2,10):
        wb_WACC.column_dimensions[letters[col]].width = 15
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.column_dimensions[letters[col]].width = 30
            wb_WACC.cell(row=row, column=col, value="Debt Breakdown")
            cell.font = boldFont

    row = 3
    for col in range(2,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            link = "http://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/ratings.htm"
            wb_WACC.cell(row=row, column=col, value='=HYPERLINK("{}", "{}")'.format(link, f"http://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/ratings.htm"))

    row = 4
    for col in range(2,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Bond Rating")
            cell.font = boldFont
        elif col == 5:
            wb_WACC.cell(row=row, column=col, value="Debt")
            cell.font = boldFont
        
    row = 5
    for col in range(2,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        cell.alignment = Alignment(horizontal="center")
        if col == 2:
            wb_WACC.cell(row=row, column=col, value=None)
            cell.font = boldFont
        elif col == 5:
            wb_WACC.cell(row=row, column=col, value="Principal")
        elif col == 6:
            wb_WACC.cell(row=row, column=col, value="Term")
        elif col == 7:
            wb_WACC.cell(row=row, column=col, value="Current Year")
        elif col == 8:
            wb_WACC.cell(row=row, column=col, value="Term Left")
        elif col == 9:
            wb_WACC.cell(row=row, column=col, value="Interest Expense")

    row = 6
    for col in range(2,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value=None)
        elif col == 5:
            wb_WACC.cell(row=row, column=col, value=None)
            cell.fill = yellowFill
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
        elif col == 6:
            wb_WACC.cell(row=row, column=col, value=None)
            cell.fill = yellowFill
        elif col == 7:
            wb_WACC.cell(row=row, column=col, value=None)
            cell.fill = yellowFill
        elif col == 8:
            wb_WACC.cell(row=row, column=col, value="=F6-G6")
            cell.fill = yellowFill
        elif col == 9:
            wb_WACC.cell(row=row, column=col, value=None)
            cell.fill = yellowFill
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO

    row = 7
    for col in range(2,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Risk Spread")
        elif col == 3:
            wb_WACC.cell(row=row, column=col, value=RISK_SPREAD)
            cell.number_format = format.FORMAT_PERCENTAGE_00
            cell.fill = yellowFill

    row = 9
    for col in range(2,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="US risk free rate normalized")
            cell.font = boldFont
        elif col == 3:
            wb_WACC.cell(row=row, column=col, value=0.04)
            cell.number_format = format.FORMAT_PERCENTAGE_00
            cell.fill = yellowFill
        elif col == 5:
            wb_WACC.cell(row=row, column=col, value="Mkt Value of Debt")
            cell.font = boldFont
        elif col == 8:
            wb_WACC.cell(row=row, column=col, value="=(I6*((1-(1/(1+C11)^H6))/C11)+E6/(1+C11)^H6)")
            cell.fill = yellowFill
            cell.font = boldFont
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
        elif col == 9:
            wb_WACC.cell(row=row, column=col, value=None)
            cell.fill = greyFill
  
    row = 11
    for col in range(2,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Pretax cost of debt")
            cell.font = boldFont
        elif col == 3:
            wb_WACC.cell(row=row, column=col, value="=C7+C9")
            cell.number_format = format.FORMAT_PERCENTAGE_00
            cell.fill = yellowFill
            
def wacc_box_two(wb_WACC):
    create_box(wb_WACC, 2, 13, 2, 4)
    row = 13
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Market Value of Equity")
            cell.font = boldFont

    row = 14
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Shares Outstanding")
        if col == 3:
            wb_WACC.cell(row=row, column=col, value="=COVER!C5")

    row = 15
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Price")
        if col == 3:
            wb_WACC.cell(row=row, column=col, value="=COVER!C2")
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO

    row = 16
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Market Cap ($m)")
            cell.font = boldFont
        if col == 3:
            wb_WACC.cell(row=row, column=col, value="=C14*C15")
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ZERO

def wacc_box_three(wb_WACC):
    create_box(wb_WACC, 2, 18, 2, 6)
    row = 18
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Debt")
        if col == 3:
            wb_WACC.cell(row=row, column=col, value='=SWITCH(I9,"thousands",H9/1000,"millions",H9,H9/1000000)')
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO

    row = 19
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Equity")
        if col == 3:
            wb_WACC.cell(row=row, column=col, value="=C16")
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ZERO

    row = 20
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Debt Weight")
        if col == 3:
            wb_WACC.cell(row=row, column=col, value="=C18/(C18+C19)")
            cell.number_format = format.FORMAT_PERCENTAGE_00

    row = 21
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Equity Weight")
        if col == 3:
            wb_WACC.cell(row=row, column=col, value="=1-C20")
            cell.number_format = format.FORMAT_PERCENTAGE_00

    row = 23
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="D/E Ratio")
        if col == 3:
            wb_WACC.cell(row=row, column=col, value="=C18/C19x`")
            cell.number_format = format.FORMAT_PERCENTAGE_00

def wacc_box_four(wb_WACC):
    create_box(wb_WACC, 2, 25, 2, 10)
    row = 25
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Beta")
            cell.font = boldFont

    row = 26
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            link = "http://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/Betas.html"
            wb_WACC.cell(row=row, column=col, value='=HYPERLINK("{}", "{}")'.format(link, f"http://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/Betas.html"))
            #wb_WACC.cell(row=row, column=col, value=link)

    row = 27
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 3:
            cell.fill = yellowFill
    
    row = 29
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="B Unlevered")
        if col == 3:
            wb_WACC.cell(row=row, column=col, value="=C27")
 
    row = 31
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="Tax Rate")
        if col == 3:
            wb_WACC.cell(row=row, column=col, value=0.35)
            cell.number_format = format.FORMAT_PERCENTAGE_00
 
    row = 32
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="D/E")
        if col == 3:
            wb_WACC.cell(row=row, column=col, value="=C23")
            cell.number_format = CUSTOM_FORMAT_PE

    row = 34
    for col in range(2,4):
        cell = wb_WACC[f"{letters[col]}{row}"]
        cell.font = boldFont
        if col == 2:
            wb_WACC.cell(row=row, column=col, value="B Unlevered")
        if col == 3:
            wb_WACC.cell(row=row, column=col, value="=C29*(1+(1-C31)*(C32))")
            cell.fill = yellowFill

def wacc_box_five(wb_WACC):
    create_box(wb_WACC, 5, 13, 5, 8)
    row = 13
    for col in range(5,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 5:
            wb_WACC.cell(row=row, column=col, value="Equity Risk Premium")
        elif col == 6:
            wb_WACC.cell(row=row, column=col, value=0.06)
            cell.number_format = format.FORMAT_PERCENTAGE_00
            cell.fill = yellowFill
        elif col == 8:
            wb_WACC.cell(row=row, column=col, value="http://pages.stern.nyu.edu/~adamodar/")

    row = 15
    for col in range(5,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 5:
            wb_WACC.cell(row=row, column=col, value="Cost of Debt")
        elif col == 6:
            wb_WACC.cell(row=row, column=col, value="=C11")
            cell.number_format = format.FORMAT_PERCENTAGE_00
    
    row = 16
    for col in range(5,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 5:
            wb_WACC.cell(row=row, column=col, value="Cost of Equity")
        elif col == 6:
            wb_WACC.cell(row=row, column=col, value="=C9+C34*F13")
            cell.number_format = format.FORMAT_PERCENTAGE_00

    row = 17
    for col in range(5,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 5:
            wb_WACC.cell(row=row, column=col, value="Debt Weight")
        elif col == 6:
            wb_WACC.cell(row=row, column=col, value="=C20")
            cell.number_format = format.FORMAT_PERCENTAGE_00

    row = 18
    for col in range(5,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 5:
            wb_WACC.cell(row=row, column=col, value="Equity Weight")
        elif col == 6:
            wb_WACC.cell(row=row, column=col, value="=C21")
            cell.number_format = format.FORMAT_PERCENTAGE_00

    row = 20
    for col in range(5,10):
        cell = wb_WACC[f"{letters[col]}{row}"]
        if col == 5:
            wb_WACC.cell(row=row, column=col, value="WACC")
            cell.fill = yellowFill
            cell.font = boldRedFont
        elif col == 6:
            wb_WACC.cell(row=row, column=col, value="=F15*F17*(1-C31)+F18*F16")
            cell.number_format = format.FORMAT_PERCENTAGE_00
            cell.fill = yellowFill
            cell.font = boldRedFont

def fill_wacc(wb_WACC):
    wacc_box_one(wb_WACC)
    wacc_box_two(wb_WACC)
    wacc_box_three(wb_WACC)
    wacc_box_four(wb_WACC)
    wacc_box_five(wb_WACC)