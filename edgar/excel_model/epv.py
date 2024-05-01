import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.styles.numbers as format

CUSTOM_FORMAT_CURRENCY_ONE = '_($* #,##0.0_);[Red]_($* (#,##0.0);_($* "-"??_)'
CUSTOM_FORMAT_CURRENCY_TWO = '_($* #,##0.00_);[Red]_($* (#,##0.00);_($* "-"??_)'
letters = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 
           15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z', 27: 'AA', 
           28: 'AB', 29: 'AC', 30: 'AD', 31: 'AE', 32: 'AF', 33: 'AG', 34: 'AH', 35: 'AI', 36: 'AJ', 37: 'AK', 38: 'AL', 39: 'AM', 
           40: 'AN', 41: 'AO', 42: 'AP', 43: 'AQ', 44: 'AR', 45: 'AS', 46: 'AT', 47: 'AU', 48: 'AV', 49: 'AW', 50: 'AX', 51: 'AY', 
           52: 'AZ', 53: 'BA', 54: 'BB', 55: 'BC', 56: 'BD', 57: 'BE', 58: 'BF', 59: 'BG', 60: 'BH', 61: 'BI', 62: 'BJ', 63: 'BK', 
           64: 'BL', 65: 'BM', 66: 'BN', 67: 'BO', 68: 'BP', 69: 'BQ', 70: 'BR', 71: 'BS', 72: 'BT', 73: 'BU', 74: 'BV', 75: 'BW', 
           76: 'BX', 77: 'BY', 78: 'BZ'
           }

retail_titles = ["Operating Income", "Depreciation Adjustment", "Depreciation", "CAPEX", "Growth CAPEX", "Option Expense", "Interest Earned on Cash",
          "Cash", "Interest Rate", "Pretax Earnings", "Tax Rate", "Taxes", "Earnings", "Earnings Power Value", "Cash", "Debt",
          "Total EV in Equity", "Shares Outstanding", "EPV/Share", "Current Share Price"]

tech_titles = ["Operating Income", "R&D Expenses", "Depreciation Adjustment", "Depreciation", "CAPEX", "Growth CAPEX", "Option Expense", "Interest Earned on Cash",
          "Cash", "Interest Rate", "Pretax Earnings", "Tax Rate", "Taxes", "Earnings", "Earnings Power Value", "Cash", "Debt",
          "Total EV in Equity", "Shares Outstanding", "EPV/Share", "Current Share Price"]

finance_titles = ["Income before income tax expense ", "Provision for Credit Losses", "Interest Expenses", "Depreciation Adjustment", "Depreciation", "CAPEX", "Growth CAPEX", "Option Expense", "Interest Earned on Cash",
          "Cash", "Interest Rate", "Pretax Earnings", "Tax Rate", "Taxes", "Earnings", "Earnings Power Value", "Cash", "Debt",
          "Total EV in Equity", "Shares Outstanding", "EPV/Share", "Current Share Price"]


depreciation_adjustment = ["Premises and Equipment", "Current Year Revenue", "Prior Year Revenue", "Change in Revenue", 
                           "Depreciation and Amortization", "CAPEX", "Growth CAPEX", "Zero-Growth CAPEX", "Depreciation Adjustment"]

boldFont = Font(name='Arial',size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
textFont  = Font(name='Arial',size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='000000')

greyFill = PatternFill(fill_type="solid", start_color='999997', end_color='999997')
yellowFill = PatternFill(fill_type="solid", start_color='ffff01', end_color='ffff01')
orangeFill = PatternFill(fill_type="solid", start_color='fde599', end_color='fde599')

greenFill = PatternFill(fill_type="solid", start_color="c6e6c1", end_color="c6e6c1")
greenerFill = PatternFill(fill_type="solid", start_color='7eab76', end_color='7eab76')
notesFill = PatternFill(fill_type="solid", start_color='fdd966', end_color='fdd966')
greenNotesFill = PatternFill(fill_type="solid", start_color='b6d7a8', end_color='b6d7a8')

depAdjFill = PatternFill(fill_type="solid", start_color="fef2cc", end_color="fef2cc")
depAdjDataFill = PatternFill(fill_type="solid", start_color="cfe2f3", end_color="cfe2f3")

titleFill = PatternFill(fill_type="solid", start_color='ebebeb', end_color='ebebeb')
thinBorder = Side(style="thin", color="000000")
thickBorder = Side(style="thick", color="000000")
noBorder = Side(style="none")

class EPVRow:
    def __init__(self):
        self.get = {}

    #def __str__(self):
    #    return ( self.row
    #    )

def EPV_titles(wb_EPV, titles, epv_row, col):
    wb_EPV.column_dimensions['B'].width = 30
    wb_EPV.cell(row=2, column=2, value="")
    cell = wb_EPV["B2"]
    cell.fill = greyFill
    cell.border = Border(left=thickBorder, top=thickBorder, right=noBorder, bottom=noBorder)
    epv_row.get["start"] = 2

    row = 3
    for i, value in enumerate(titles):
        wb_EPV.cell(row=row, column=col, value=value)
        cell = wb_EPV[f"B{row}"]
        cell.font = boldFont
        cell.fill = titleFill
        if value in epv_row.get:
            epv_row.get[value] = [epv_row.get[value], row]
        else:
            epv_row.get[value] = row
        if value == "Depreciation Adjustment":
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thinBorder)
        elif value == "Growth CAPEX":
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thinBorder)
        elif value == "Shares Outstanding":
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thinBorder)
        elif value == "EPV/share":
            cell.border = Border(left=thickBorder, top=thinBorder, right=noBorder, bottom=noBorder)
        elif value == "Current Share Price":
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thickBorder)
            epv_row.get["end"] = row
        else:
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        row += 1
    
def dep_adj_titles(wb_EPV, epv_row, col):
    row = epv_row.get["end"] + 2
    for i, value in enumerate(depreciation_adjustment):
        wb_EPV.cell(row=row, column=col, value=value)
        cell = wb_EPV[f"B{row}"]
        cell.font = textFont
        cell.fill = depAdjFill

        if value in epv_row.get:
            epv_row.get[value] = [epv_row.get[value], row]
        else:
            epv_row.get[value] = row
        if (value == "Premises and Equipment"):
            cell.border = Border(left=thickBorder, top=thickBorder, right=noBorder, bottom=noBorder)
        elif (value == "Zero-Growth CAPEX"):
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thinBorder)
        elif (value == "Depreciation Adjustment"):
            cell.font = boldFont
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thickBorder)
        else:
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        row += 1

def fill_EPV_data(wb_EPV, epv_row, col, titles, epv_info, iSharesRow, industry):
    col = 3
    for y, date in enumerate(epv_info):
        wb_EPV.column_dimensions[f"{letters[col]}"].width = 25
        row = 2
        year = date.split('-')[0]
        wb_EPV.cell(row=row, column=col, value=year)
        cell = wb_EPV[f"{letters[col]}2"]
        cell.fill = greenerFill
        cell.alignment = Alignment(horizontal="center")
        cell.font = boldFont
        cell.border = Border(left=thinBorder, top=thickBorder, right=thinBorder, bottom=noBorder)
        row += 1
        
        for i, epv_title in enumerate(titles):
            value = epv_info[date][epv_title] if epv_title in epv_info[date] else 0
            # Debt needs to be negative
            if epv_title == "Debt":
                if value is not None:
                    value = -abs(value)
            cell = wb_EPV[f"{letters[col]}{row}"]
            cell.fill = greenFill
            # Operating Income
            if epv_title == "Operating Income":
                wb_EPV.cell(row=row, column=col, value=value)
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Income Before Income Tax Expense
            if epv_title == "Income Before Income Tax Expense":
                wb_EPV.cell(row=row, column=col, value=value)
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Provision for Credit Losses
            if epv_title == "Provision for Credit Losses":
                wb_EPV.cell(row=row, column=col, value=value)
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE            
            # Interest Expenses
            if epv_title == "Interest Expenses ":
                wb_EPV.cell(row=row, column=col, value=value)
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE            
            # R&D Expenses
            elif epv_title == "R&D Expenses":
                wb_EPV.cell(row=row, column=col, value=f"={value}/15")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Depreciation Adjustment
            elif epv_title == "Depreciation Adjustment":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['Depreciation Adjustment'][1]}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=thinBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Depreciation
            elif epv_title == "Depreciation":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['Depreciation and Amortization']}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # CAPEX
            elif epv_title == "CAPEX":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['CAPEX'][1]}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)      
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Growth CAPEX
            elif epv_title == "Growth CAPEX":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['Growth CAPEX'][1]}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=thinBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Option Expense
            elif epv_title == "Option Expense":
                wb_EPV.cell(row=row, column=col, value=value)
                cell.border = Border(left=thinBorder, top=thinBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Interest Earned on Cash
            elif epv_title == "Interest Earned on Cash":
                wb_EPV.cell(row=row, column=col, value=f"=-{letters[col]}{epv_row.get['Cash'][0]}*{letters[col]}{epv_row.get['Interest Rate']}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Cash
            elif epv_title == "Cash":
                if row == epv_row.get["Interest Earned on Cash"]+1:
                    wb_EPV.cell(row=row, column=col, value=value)
                else:
                    wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['Cash'][0]}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Interest Rate
            elif epv_title == "Interest Rate":
                int_year = int(year)
                if int_year >= 2023:
                    year_val = 0.06
                elif int_year == 2022:
                    year_val = 0.04
                else:
                    year_val = 0.02
                wb_EPV.cell(row=row, column=col, value=year_val)
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = format.FORMAT_PERCENTAGE_00
            # Pretax Earnings
            elif epv_title == "Pretax Earnings":
                income_row = epv_row.get['Income before income tax expense '] if industry == "finance" else epv_row.get['Operating Income']
                wb_EPV.cell(row=row, column=col, value=f"=sum({letters[col]}{income_row}:{letters[col]}{epv_row.get['Depreciation Adjustment'][0]})+{letters[col]}{epv_row.get['Interest Earned on Cash']}-{letters[col]}{epv_row.get['Option Expense']}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Tax Rate
            elif epv_title == "Tax Rate":
                wb_EPV.cell(row=row, column=col, value=0.21)
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = format.FORMAT_PERCENTAGE_00
            # Taxes
            elif epv_title == "Taxes":
                wb_EPV.cell(row=row, column=col, value=f"=-{letters[col]}{epv_row.get['Pretax Earnings']}*{letters[col]}{epv_row.get['Tax Rate']}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Earnings
            elif epv_title == "Earnings":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['Pretax Earnings']}+{letters[col]}{epv_row.get['Taxes']}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Earnings Power Value
            elif epv_title == "Earnings Power Value":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['Earnings']}/WACC!F20")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Debt
            elif epv_title == "Debt":
                wb_EPV.cell(row=row, column=col, value=value)
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE            
            # Total EV in Equity
            elif epv_title == "Total EV in Equity":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['Earnings Power Value']}+{letters[col]}{epv_row.get['Cash'][1]}+{letters[col]}{epv_row.get['Debt']}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            # Shares Outstanding
            elif epv_title == "Shares Outstanding":
                iNAVSharesCol = 3*(y+3) - 6
                wb_EPV.cell(row=row, column=col, value=f"=NAV!{letters[iNAVSharesCol]}{iSharesRow}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=thinBorder)
                cell.number_format = format.FORMAT_NUMBER_COMMA_SEPARATED1
            # EPV/Share
            elif epv_title == "EPV/Share":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['Total EV in Equity']}/{letters[col]}{epv_row.get['Shares Outstanding']}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
                cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
                cell.fill = yellowFill
                cell.font = boldFont
            # Current Share Price
            elif epv_title == "Current Share Price":
                wb_EPV.cell(row=row, column=col, value="=COVER!C2")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=thickBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
                cell.fill = yellowFill
                cell.font = boldFont
            else:
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            row += 1 
        col += 1
    return [letters[col-1], epv_row.get['EPV/Share']]

def fill_dep_adj_data(wb_EPV, epv_row, col, epv_info):
    col = 3
    for y, date in enumerate(epv_info):
        row = epv_row.get["Premises and Equipment"]
        for i, epv_title in enumerate(depreciation_adjustment):
            value = epv_info[date][epv_title] if epv_title in epv_info[date] else None
            # CAPEX needs to always positive
            if epv_title == "CAPEX":
                value = abs(value) if value else 0
            wb_EPV.cell(row=row, column=col, value=value)
            cell = wb_EPV[f"{letters[col]}{row}"]
            cell.font = textFont
            cell.fill = depAdjDataFill
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            
            # Premises and Equipment
            if epv_title == "Premises and Equipment":
                cell.border = Border(left=thinBorder, top=thickBorder, right=thinBorder, bottom=noBorder)
            # Prior Year Revenue
            elif epv_title == "Prior Year Revenue":
                if letters[col] == 'C':
                    wb_EPV.cell(row=row, column=col, value=None)
                else:
                    wb_EPV.cell(row=row, column=col, value=f"={letters[col-1]}{epv_row.get['Current Year Revenue']}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)                
            # Change in Revenue
            elif epv_title == "Change in Revenue":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['Current Year Revenue']}-{letters[col]}{epv_row.get['Prior Year Revenue']}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
            # Growth CAPEX:
            elif epv_title == "Growth CAPEX":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['Premises and Equipment']}/{letters[col]}{epv_row.get['Current Year Revenue']}*{letters[col]}{epv_row.get['Change in Revenue']}")
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
            # Zero-Growth CAPEX
            elif epv_title == "Zero-Growth CAPEX":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['CAPEX'][1]}-{letters[col]}{epv_row.get['Growth CAPEX'][1]}")
                cell.border = Border(left=thinBorder, top=noBorder, right=thinBorder, bottom=thinBorder)
            # Depreciation Adjustment
            elif epv_title == "Depreciation Adjustment":
                wb_EPV.cell(row=row, column=col, value=f"={letters[col]}{epv_row.get['Depreciation and Amortization']}-{letters[col]}{epv_row.get['Zero-Growth CAPEX']}")
                cell.font = boldFont
                cell.border = Border(left=thinBorder, top=noBorder, right=thinBorder, bottom=thickBorder)
            else:
                cell.border = Border(left=thinBorder, top=noBorder, right=noBorder, bottom=noBorder)
                cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            row += 1
        col += 1

def fill_notes(wb_EPV, epv_row, col):
    row = epv_row.get["start"]
    cell = wb_EPV[f"{letters[col]}{row}"]
    wb_EPV.cell(row=row, column=col, value="Notes")
    wb_EPV.column_dimensions[f"{letters[col]}"].width = 15
    cell.font = boldFont
    cell.alignment = Alignment(horizontal="center")
    cell.fill = notesFill
    cell.border = Border(left=thinBorder, top=thickBorder, right=thickBorder, bottom=noBorder)

    # EPV notes
    row = 3
    for _ in range(3, epv_row.get["Current Share Price"]+1):
        wb_EPV.cell(row=row, column=col, value=None)
        cell = wb_EPV[f"{letters[col]}{row}"]
        cell.fill = orangeFill  

        if row == epv_row.get["Current Share Price"]:
            cell.border = Border(left=thinBorder, top=noBorder, right=thickBorder, bottom=thickBorder)
        else:
            cell.border = Border(left=thinBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        row += 1
    
    # DA Notes
    row = epv_row.get["Premises and Equipment"]
    for _ in range(epv_row.get["Premises and Equipment"], epv_row.get["Depreciation Adjustment"][1]+1):
        wb_EPV.cell(row=row, column=col, value=None)
        cell = wb_EPV[f"{letters[col]}{row}"]
        cell.fill = greenNotesFill
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=thinBorder, top=noBorder, right=thickBorder, bottom=noBorder)

        if row == epv_row.get["Premises and Equipment"]:
            wb_EPV.cell(row=row, column=col, value="(a)")
            cell.border = Border(left=thinBorder, top=thickBorder, right=thickBorder, bottom=noBorder)
        elif row == epv_row.get["Current Year Revenue"]:
            wb_EPV.cell(row=row, column=col, value="(b)")
        elif row == epv_row.get["Prior Year Revenue"]:
            wb_EPV.cell(row=row, column=col, value="(c)")
        elif row == epv_row.get["Change in Revenue"]:
            wb_EPV.cell(row=row, column=col, value="(d) = (b) - (c)")
        elif row == epv_row.get["Depreciation and Amortization"]:
            wb_EPV.cell(row=row, column=col, value="(e)")
        elif row == epv_row.get["CAPEX"]:
            wb_EPV.cell(row=row, column=col, value="(f)")
        elif row == epv_row.get["Growth CAPEX"]:
            wb_EPV.cell(row=row, column=col, value="(g) = (a)/(b) x (d)")
        elif row == epv_row.get["Zero-Growth CAPEX"]:
            cell.border = Border(left=thinBorder, top=noBorder, right=thickBorder, bottom=thinBorder)
            wb_EPV.cell(row=row, column=col, value="(h) = (f) - (g)")
        elif row == epv_row.get["Depreciation Adjustment"][1]:
            cell.border = Border(left=thinBorder, top=thinBorder, right=thickBorder, bottom=thickBorder)
            wb_EPV.cell(row=row, column=col, value="(i) = (e) - (h)")
        row += 1

def fill_epv(wb_EPV, industry, epv_info, iSharesRow):
    wb_EPV.column_dimensions['A'].width = 2
    row, col = 3, 2
    epv_row = EPVRow()
    
    titles = []
    if industry == "tech": titles = tech_titles
    elif industry == "retail": titles = retail_titles
    elif industry == "finance": titles = finance_titles

    EPV_titles(wb_EPV, titles, epv_row, col)
    dep_adj_titles(wb_EPV, epv_row, col)
    col = 3
    iEPVPriceCoord = fill_EPV_data(wb_EPV, epv_row, col, titles, epv_info, iSharesRow, industry)
    fill_dep_adj_data(wb_EPV, epv_row, col, epv_info)
    col = 3+len(epv_info)
    fill_notes(wb_EPV, epv_row, col)
    return epv_row, iEPVPriceCoord

if __name__ == "__main__":
    ticker = "DOMI"
    path = f"./EPV/{ticker}.xlsx"	

    wb = xl.Workbook()
    wb.create_sheet("EPV")

    row, col = 3, 2
    wb_EPV = wb["EPV"]
    EPV_titles(wb_EPV, row, col)
    row = 24
    dep_adj_titles(wb_EPV, row, col)
    row, col = 3, 3
    fill_EPV_data(wb_EPV, row, col,)
    row = 24
    fill_dep_adj_data(wb_EPV, row, col)
    wb.save(path)