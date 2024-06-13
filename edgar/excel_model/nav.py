import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.styles.numbers as format

CUSTOM_FORMAT_CURRENCY_ONE = '_($* #,##0.0_);[Red]_($* (#,##0.0);_($* "-"??_)'
CUSTOM_FORMAT_CURRENCY_TWO = '_($* #,##0.00_);[Red]_($* (#,##0.00);_($* "-"??_)'
letters = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 
           15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z', 27: 'AA', 
           28: 'AB', 29: 'AC', 30: 'AD', 31: 'AE', 32: 'AF', 33: 'AG', 34: 'AH', 35: 'AI', 36: 'AJ', 37: 'AK', 38: 'AL', 39: 'AM', 
           40: 'AN', 41: 'AO', 42: 'AP', 43: 'AQ', 44: 'AR', 45: 'AS', 46: 'AT', 47: 'AU', 48: 'AV', 49: 'AW', 50: 'AX', 51: 'AY', 
           52: 'AZ', 53: 'BA', 54: 'BB', 55: 'BC', 56: 'BD', 57: 'BE', 58: 'BF', 59: 'BG', 60: 'BH', 61: 'BI', 62: 'BJ', 63: 'BK', 
           64: 'BL', 65: 'BM', 66: 'BN', 67: 'BO', 68: 'BP', 69: 'BQ', 70: 'BR', 71: 'BS', 72: 'BT', 73: 'BU', 74: 'BV', 75: 'BW', 
           76: 'BX', 77: 'BY', 78: 'BZ'
           }

boldFont = Font(name='Arial',size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
titleFont = Font(name='Arial',size=14, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
headingFont = Font(name='Arial',size=10, bold=True, italic=False, vertAlign=None, underline='single', strike=False, color='000000')
textFont  = Font(name='Arial',size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='000000')

greyFill = PatternFill(fill_type="solid", start_color='ebebeb', end_color='ebebeb')
greyerFill = PatternFill(fill_type="solid", start_color='999997', end_color='999997')
purpleFill = PatternFill(fill_type="solid", start_color='d9d2e9', end_color='d9d2e9')
purplerFill = PatternFill(fill_type="solid", start_color='b4a7d7', end_color='b4a7d7')
blueFill = PatternFill(fill_type="solid", start_color='c9dbf8', end_color='c9dbf8')
bluerFill = PatternFill(fill_type="solid", start_color='a4c2f4', end_color='a4c2f4')
greenFill = PatternFill(fill_type="solid", start_color="c6e6c1", end_color="c6e6c1")
greenerFill = PatternFill(fill_type="solid", start_color='93c47d', end_color='93c47d')
yellowishFill = PatternFill(fill_type="solid", start_color='fff2cc', end_color='fff2cc')

depAdjFill = PatternFill(fill_type="solid", start_color="fef2cc", end_color="fef2cc")
depAdjDataFill = PatternFill(fill_type="solid", start_color="cfe2f3", end_color="cfe2f3")

titleFill = PatternFill(fill_type="solid", start_color='ebebeb', end_color='ebebeb')
thinBorder = Side(style="thin", color="000000")
thickBorder = Side(style="thick", color="000000")
noBorder = Side(style="none")

class AssetRow:
    def __init__(self):
        self.start = 2
        self.current_asset = None
        self.non_current_assert = None
        self.PPE = None
        self.SGA = None
        self.total_asset = None
        self.end = None
    
    def __str__(self):
        return (f"start = {self.start}\n" + 
                f"current asset = {self.current_asset}\n" + 
                f"non current assert = {self.non_current_assert}\n" +
                f"PPE = {self.PPE}\n" +
                f"SG&A = {self.SGA}\n" +
                f"Total Assets = {self.total_asset}\n" +
                f"End = {self.end}\n"
        )

class LiabilityRow:
    def __init__(self):
        self.start = None
        self.current_liability = None
        self.non_current_liability = None
        self.contract = None
        self.options = None
        self.total_liability = None
        self.end = None

    def __str__(self):
        return (f"start = {self.start}\n" + 
                f"current liability = {self.current_liability}\n" + 
                f"non current liability = {self.non_current_liability}\n" +
                f"Contractual Obli = {self.contract}\n" +
                f"Options = {self.options}\n" +
                f"Total Liability = {self.total_liability}\n" +
                f"End = {self.end}\n"
        )

class SummaryRow:
    def __init__(self):
        self.start = None
        self.NAV = None
        self.shares = None
        self.NAV_price = None
        self.current_price = None
        self.end = None

    def __str__(self):
        return (f"Start = {self.start}\n" + 
                f"NAV = {self.NAV}\n" + 
                f"Shares Outstanding = {self.shares}\n" + 
                f"NAV Price Per Share = {self.NAV_price}\n" +
                f"Current Price = {self.current_price}\n"
        )

PPE = ["us-gaap:PropertyPlantAndEquipmentNet"]

def is_PPE(tag):
    return any(tag == PPE_tag for PPE_tag in PPE)

def NAV_assets_titiles(wb_NAV, assets_info, all_dates, k_dates, asset_row):
    wb_NAV.column_dimensions['B'].width = 70
    # Title Assets
    wb_NAV.cell(row=asset_row.start, column=2, value="Assets")
    cell = wb_NAV["B2"]
    cell.fill = greyerFill
    cell.font = titleFont
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(left=thickBorder, top=thickBorder, right=thickBorder, bottom=thinBorder)

    row = asset_row.start+1
    # Start of Current Assets (for now, it will also add data for all Non-Current Assets)
    wb_NAV.cell(row=row, column=2, value="Current Assets")
    cell = wb_NAV["B3"]
    cell.fill = greyFill
    cell.font = headingFont
    cell.border = Border(left=thickBorder, top=thinBorder, right=thickBorder, bottom=noBorder) 

    row += 1
    col = 2
    asset_row.current_asset = row
    PPE_tag_info = None
    for asset_tag_info in assets_info:
        # Skip cell if cell is Total Non-Current Assets or Total Assets
        if asset_tag_info["Tag"] == "us-gaap:Assets" or asset_tag_info["Tag"] == "us-gaap:AssetsNoncurrent":
            continue
        elif is_PPE(asset_tag_info["Tag"]):
            PPE_tag_info = asset_tag_info
            continue

        cell = wb_NAV[f"{letters[col]}{row}"]
        # Cell value to Non-Current Asset as it is start of it
        if asset_tag_info["Tag"] == "us-gaap:AssetsCurrent":
            wb_NAV.cell(row=row, column=2, value="Non-Current Assets")
            cell.font = headingFont
        else:
            wb_NAV.cell(row=row, column=2, value=f"{asset_tag_info['Text']}")
        cell.fill = greyFill
        cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        row += 1

    # PPE
    asset_row.PPE = row
    wb_NAV.cell(row=row, column=2, value="PPE")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greyFill
    cell.font = headingFont
    cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
    row += 1
    # Add PPE data
    if PPE_tag_info is not None:
        wb_NAV.cell(row=row, column=2, value=PPE_tag_info["Text"])
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = greyFill
        cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        row += 1
    for _ in range(5):
        wb_NAV.cell(row=row, column=2, value=None)
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = greyFill
        cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        row += 1
    
    asset_row.SGA = row
    # Goodwill
    wb_NAV.cell(row=row, column=2, value="Goodwill")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greyFill
    cell.font = headingFont
    cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
    row += 1

    year = all_dates[0].split('-')[0]
    cur_year = int(year)-2

    # If 1, it means the most recent filing is a 10-Q so we need to add an extra row of SG&A for new fiscal year.
    i_q_recent = 1 if all_dates[-1] != k_dates[-1] else 0
    for i in range(len(k_dates)+2+i_q_recent):
        # If the date is a 10-Q date, we don't want to add row
        wb_NAV.cell(row=row, column=2, value=f"{cur_year} SG&A")
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = greyFill
        if i == 0:
            cell.border = Border(left=thickBorder, top=thinBorder, right=thickBorder, bottom=noBorder)
        else:
            cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        cur_year += 1
        row += 1

    wb_NAV.cell(row=row, column=2, value="Average SG&A")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greyFill
    cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=thinBorder)
    row += 1

    # Total Assets
    asset_row.total_asset = row
    wb_NAV.cell(row=row, column=2, value="Total Assets")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greyFill
    cell.font = boldFont
    cell.border = Border(left=thickBorder, top=thinBorder, right=thickBorder, bottom=noBorder)
    row += 1
    
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greyFill
    cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
    row += 1
    asset_row.end = row

def NAV_liabilities_titiles(wb_NAV, liabilities_info, row, liability_row):
    # Title Liabilities
    liability_row.start = row
    wb_NAV.cell(row=row, column=2, value="Liabilities")
    cell = wb_NAV[f"B{row}"]
    cell.fill = greyerFill
    cell.font = titleFont
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(left=thickBorder, top=thinBorder, right=thickBorder, bottom=thinBorder)
    row += 1

    # Start of Current Liabilites (for now, it will also add data for all Non-Current Liabilities)
    liability_row.current_liability = row
    wb_NAV.cell(row=row, column=2, value="Current Liabilities")
    cell = wb_NAV[f"B{row}"]
    cell.fill = greyFill
    cell.font = headingFont
    cell.border = Border(left=thickBorder, top=thinBorder, right=thickBorder, bottom=noBorder) 
    row += 1

    col = 2
    for liability_tag_info in liabilities_info:
        # Skip if Total Non-Current Liabilities or Total Liabilities
        if liability_tag_info["Tag"] == "us-gaap:Liabilities" or liability_tag_info["Tag"] == "us-gaap:LiabilitiesNoncurrent":
            continue

        cell = wb_NAV[f"{letters[col]}{row}"]
        # Cell value to Non-Current Asset as it is start of it
        if liability_tag_info["Tag"] == "us-gaap:LiabilitiesCurrent":
            wb_NAV.cell(row=row, column=2, value="Non-Current Liabilities")
            cell.font = headingFont
        else:
            wb_NAV.cell(row=row, column=2, value=f"{liability_tag_info['Text']}")
        cell.fill = greyFill
        cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        row += 1
    
    # Titles for PPE and Goodwill

    # Contractual Obligations & Off-Balance Sheet Arrangements
    liability_row.contract = row
    wb_NAV.cell(row=row, column=2, value="Contractual Obligations & Off-Balance Sheet Arrangements")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greyFill
    cell.font = headingFont
    cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
    row += 1
    for _ in range(3):
        wb_NAV.cell(row=row, column=2, value=None)
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = greyFill
        cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        row += 1
    
    # Future Option Grants
    liability_row.options = row
    wb_NAV.cell(row=row, column=2, value="Future Option Grants")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greyFill
    cell.font = headingFont
    cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
    row += 1
    for i in range(6):
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = greyFill
        if i % 2 == 1:
            wb_NAV.cell(row=row, column=2, value="Weighted Average Exercise Price")
        if i == 0:
            cell.border = Border(left=thickBorder, top=thinBorder, right=thickBorder, bottom=noBorder)
        else:
            cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        row += 1

    wb_NAV.cell(row=row, column=2, value="Esitmated Option Liability")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greyFill
    cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=thinBorder)
    row += 1

    # Total Liabilities
    liability_row.total_liability = row
    wb_NAV.cell(row=row, column=2, value="Total Liabilities")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greyFill
    cell.font = boldFont
    cell.border = Border(left=thickBorder, top=thinBorder, right=thickBorder, bottom=thinBorder)
    row += 1

    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.border = Border(left=thickBorder, top=thinBorder, right=thickBorder, bottom=noBorder)
    row += 1
    liability_row.end = row
    
def NAV_summary_titles(wb_NAV, row, summary_row):
    summary_row.start = row
    col = 2
    # Net Asset Value (NAV)
    summary_row.NAV = row
    wb_NAV.cell(row=row, column=2, value="Net Asset Value (NAV)")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.font = boldFont
    cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
    row += 1
    # Shares Oustanding
    summary_row.shares = row
    wb_NAV.cell(row=row, column=2, value="Shares Oustanding")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.font = boldFont
    cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
    row += 1
    # NAV Share Price
    summary_row.NAV_price = row
    wb_NAV.cell(row=row, column=2, value="NAV Share Price")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.font = boldFont
    cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=noBorder)
    row += 1
    # Current Share Price
    summary_row.current_price = row
    wb_NAV.cell(row=row, column=2, value="Current Share Price")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.font = boldFont
    cell.border = Border(left=thickBorder, top=noBorder, right=thickBorder, bottom=thickBorder)
    summary_row.end = row

def NAV_assets_data(wb_NAV, assets_info, col, asset_row, date, all_dates, k_dates, iSGA, iNumYears, i_quarter, SGA_list):
    wb_NAV.column_dimensions[letters[col]].width = 15
    year = date.split('-')[0]
    # Title Assets
    if date in k_dates:
        wb_NAV.cell(row=2, column=col, value=f"FY {year}")
    else:
        wb_NAV.cell(row=2, column=col, value=f"Q{i_quarter} {year}")

    cell = wb_NAV[f"{letters[col]}2"]
    cell.fill = purplerFill
    cell.font = boldFont
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(left=thickBorder, top=thickBorder, right=noBorder, bottom=thinBorder)

    # Start of Current Assets (for now, it will also add data for all Non-Current Assets)
    cell = wb_NAV[f"{letters[col]}3"]
    cell.fill = purpleFill
    cell.border = Border(left=thickBorder, top=thinBorder, right=noBorder, bottom=noBorder) 

    row = 4
    PPE_tag_info = None
    for asset_tag_info in assets_info:
        # Skip cell if cell is Total Non-Current Assets or Total Assets
        if asset_tag_info["Tag"] == "us-gaap:Assets" or asset_tag_info["Tag"] == "us-gaap:AssetsNoncurrent":
            continue
        elif is_PPE(asset_tag_info["Tag"]):
            PPE_tag_info = asset_tag_info
            continue

        value = asset_tag_info[date] if date in asset_tag_info else 0
        # Don't add data for Total Current Assets
        if asset_tag_info["Tag"] == "us-gaap:AssetsCurrent":
            value = None

        wb_NAV.cell(row=row, column=col, value=value)
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = purpleFill
        cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
        row += 1
    
    for _ in range(row, asset_row.SGA):
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = purpleFill
        cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE

        # If the row is PPE data, add that
        if row == asset_row.PPE+1 and PPE_tag_info is not None:
            value = PPE_tag_info[date] if date in PPE_tag_info else 0
            wb_NAV.cell(row=row, column=col, value=value)
        row += 1
    
    # SG&A and Total Asset
    j = 0
    for _ in range(row, asset_row.end):
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = purpleFill
        if row > asset_row.SGA and row < asset_row.total_asset-1:
            if iSGA == 0:
                # We don't have data for teh old SG&A in the case were the oldest filing is 10-Q so just put in 0
                value = SGA_list[j] if j < len(SGA_list) else 0
                wb_NAV.cell(row=row, column=col, value=value)
                j += 1
            else:
                wb_NAV.cell(row=row, column=col, value=f"={letters[col-3]}{row}")
        # Total Asset
        if row == asset_row.end-2:
            wb_NAV.cell(row=row, column=col, value=f"=SUM({letters[col]}{asset_row.current_asset}:{letters[col]}{asset_row.PPE+1})+{letters[col]}{asset_row.total_asset-1}")
            cell.border = Border(left=thickBorder, top=thinBorder, right=noBorder, bottom=noBorder)
        # Fill in the AVERAGE formula for SG&A
        elif row == asset_row.total_asset-1:
            iAvgRow = asset_row.total_asset - 1
            cell = wb_NAV[f"{letters[col]}{iAvgRow}"]
            iStartAvgRow = iAvgRow - 2 - iNumYears + iSGA
            iEndAvgRow = iAvgRow - iNumYears + iSGA
            # Go back an extra year if most recent filing is 10-Q since we are adding a new SG&A row for the new fiscal year.
            if all_dates[-1] != k_dates[-1]:
                iStartAvgRow -= 1
                iEndAvgRow -= 1
            wb_NAV.cell(row=iAvgRow, column=col, value=f"=AVERAGE({letters[col]}{iStartAvgRow}:{letters[col]}{iEndAvgRow})")
        else:
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
        row += 1
    
    # Total Asset Calculation (unadjusted)

def NAV_liabilities_data(wb_NAV, liabilities_info, col, liability_row, date):
    row = liability_row.start
    # Title Liabilities
    year = str(date.split('-')[0])
    wb_NAV.cell(row=row, column=col, value=f"={letters[col]}{2}")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = purplerFill
    cell.font = boldFont
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(left=thickBorder, top=thinBorder, right=noBorder, bottom=thinBorder)
    row += 1

    # Start of Current Liabilities (for now, it will also add data for all Non-Current Liabilities)
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = purpleFill
    cell.border = Border(left=thickBorder, top=thinBorder, right=noBorder, bottom=noBorder) 
    row += 1

    for liability_tag_info in liabilities_info:
        # Skip if Total Non-Current Liabilities or Total Liabilities
        if liability_tag_info["Tag"] == "us-gaap:Liabilities" or liability_tag_info["Tag"] == "us-gaap:LiabilitiesNoncurrent":
            continue
        
        value = liability_tag_info[date] if date in liability_tag_info else 0
        # Don't add data for Total Current Liabilities
        if liability_tag_info["Tag"] == "us-gaap:LiabilitiesCurrent":
            value = None

        wb_NAV.cell(row=row, column=col, value=value)
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = purpleFill
        cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
        row += 1
    
    # Rows up till Options
    for _ in range(row, liability_row.options+1):
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = purpleFill        
        cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
        row += 1
    
    # Options
    i = 0
    for _ in range(row, liability_row.total_liability):
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = purpleFill        
        cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        if i%2 == 0:
            cell.number_format = format.BUILTIN_FORMATS[3]
        else:
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
        
        if row == liability_row.total_liability-1:
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
            wb_NAV.cell(row=row, column=col, value=f"={letters[col]}{row-6}*{letters[col]}{row-5}+{letters[col]}{row-4}*{letters[col]}{row-3}+{letters[col]}{row-2}*{letters[col]}{row-1}")

        row += 1
        i += 1

    # Total Liability
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = purpleFill
    wb_NAV.cell(row=row, column=col, value=f"=SUM({letters[col]}{liability_row.current_liability+1}:{letters[col]}{liability_row.options-1})+{letters[col]}{liability_row.total_liability-1}")
    cell.border = Border(left=thickBorder, top=thinBorder, right=noBorder, bottom=thinBorder)
    cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE

def NAV_summary_data(wb_NAV, col, total_asset_row, start_row, end_row, shares):
    row = start_row
    for _ in range(start_row, end_row+1):
        cell = wb_NAV[f"{letters[col]}{row}"]   
        # NAV Share Price
        if row == end_row-1:
            wb_NAV.cell(row=row, column=col, value=f"={letters[col]}{row-2}/{letters[col]}{row-1}")
            cell.fill = yellowishFill
            cell.font = boldFont
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
        # Current Share Price
        elif row == end_row:
            wb_NAV.cell(row=row, column=col, value="=COVER!C2")
            cell.fill = yellowishFill
            cell.font = boldFont
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=thickBorder)
            cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
        else:
            # Net Asset Value (NAV)
            if row == end_row-3:
                wb_NAV.cell(row=row, column=col, value=f"={letters[col+2]}{total_asset_row}+{letters[col+2]}{row-2}")
                cell.font = boldFont
                cell.number_format = CUSTOM_FORMAT_CURRENCY_TWO
            #   Shares Oustanding
            elif row == end_row-2:
                shares_string = f'''=SWITCH(,"thousands",{shares}/1000,"millions",{shares}/1000000,{shares})'''
                shares_string = shares
                wb_NAV.cell(row=row, column=col, value=shares)
                cell.number_format = format.FORMAT_NUMBER_COMMA_SEPARATED1
            cell.border = Border(left=thickBorder, top=noBorder, right=noBorder, bottom=noBorder)
        row += 1

def NAV_asset_adjustment(wb_NAV, col, assets_info, asset_row, bIsFirstCol):
    wb_NAV.column_dimensions[letters[col]].width = 15
    row = asset_row.start

    # Adjusted
    wb_NAV.cell(row=row, column=col, value="Adjusted")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = bluerFill
    cell.font = boldFont
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(left=noBorder, top=thickBorder, right=noBorder, bottom=thinBorder)
    row += 1

    # Start of Current Assets (for now, it will also add data for all Non-Current Assets)
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = blueFill
    cell.border = Border(left=noBorder, top=thinBorder, right=noBorder, bottom=noBorder) 
    row += 1

    PPE_tag_info = None
    # Adjustments for Asset Data
    for asset_tag_info in assets_info:
        # Skip cell if cell is Total Non-Current Assets or Total Assets
        if asset_tag_info["Tag"] == "us-gaap:Assets" or asset_tag_info["Tag"] == "us-gaap:AssetsNoncurrent":
            continue
        # Add adjustment for PPE
        elif is_PPE(asset_tag_info["Tag"]):
            PPE_tag_info = asset_tag_info
            continue

        if bIsFirstCol: value = 1
        else: value = f"={letters[col-3]}{row}"
        # Don't add adjustment for Total Current Asset Cell
        if asset_tag_info["Tag"] == "us-gaap:AssetsCurrent":
            value = None
        

        wb_NAV.cell(row=row, column=col, value=value)
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = blueFill
        cell.border = Border(left=noBorder, top=noBorder, right=noBorder, bottom=noBorder)
        cell.number_format = format.FORMAT_PERCENTAGE
        row += 1 

    for _ in range(row, asset_row.end):
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = blueFill

        # Add PPE adjustment percentage.
        # TODO: Change to formula
        if row == asset_row.PPE+1 and PPE_tag_info is not None:
            wb_NAV.cell(row=row, column=col, value=1)
            cell.number_format = format.FORMAT_PERCENTAGE

        # Add SG&A percentage adjustment
        elif row == asset_row.end-3:
            if bIsFirstCol:
                wb_NAV.cell(row=row, column=col, value=1)
            else:
                wb_NAV.cell(row=row, column=col, value=f"={letters[col-3]}{row}")
            cell.number_format = format.FORMAT_PERCENTAGE
        if row == asset_row.end-2:
            cell.border = Border(left=noBorder, top=thinBorder, right=noBorder, bottom=noBorder)
        else:
            cell.border = Border(left=noBorder, top=noBorder, right=noBorder, bottom=noBorder)
        row += 1

def NAV_liability_adjustment(wb_NAV, col, liabilities_info, liability_row, bIsFirstCol):
    row = liability_row.start
    # Adjusted
    wb_NAV.cell(row=row, column=col, value="Adjusted")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = bluerFill
    cell.font = boldFont
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(left=noBorder, top=thinBorder, right=noBorder, bottom=thinBorder)
    row += 1

    # Start of Current Liability (for now, it will also add data for all Non-Current Assets)
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = blueFill
    cell.border = Border(left=noBorder, top=thinBorder, right=noBorder, bottom=noBorder) 
    row += 1

    # Adjustments for Liability Data
    for liability_tag_info in liabilities_info:
        # Skip if Total Non-Current Liabilities or Total Liabilities
        if liability_tag_info["Tag"] == "us-gaap:Liabilities" or liability_tag_info["Tag"] == "us-gaap:LiabilitiesNoncurrent":
            continue

        if bIsFirstCol: value = -1
        else: value = f"={letters[col-3]}{row}"

        # Don't add data for Total Current Assets
        if liability_tag_info["Tag"] == "us-gaap:LiabilitiesCurrent":
            value = None

        wb_NAV.cell(row=row, column=col, value=value)
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = blueFill
        cell.border = Border(left=noBorder, top=noBorder, right=noBorder, bottom=noBorder)
        cell.number_format = format.FORMAT_PERCENTAGE
        row += 1 

    for _ in range(row, liability_row.end):
        cell = wb_NAV[f"{letters[col]}{row}"]
        if row == liability_row.end-1:
            break

        cell.fill = blueFill
        # Total Liabilities
        if row == liability_row.end-2:
            cell.border = Border(left=noBorder, top=thinBorder, right=noBorder, bottom=thinBorder)
        else:
            # Estimated Option Liability
            if row == liability_row.end-3:
                wb_NAV.cell(row=row, column=col, value=-1)
                cell.number_format = format.FORMAT_PERCENTAGE
            cell.border = Border(left=noBorder, top=noBorder, right=noBorder, bottom=noBorder)
        row += 1
   
def NAV_asset_adjusted_data(wb_NAV, col, assets_info, asset_row):
    wb_NAV.column_dimensions[letters[col]].width = 15
    row = asset_row.start

    # Adjusted
    wb_NAV.cell(row=row, column=col, value="Value")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greenerFill
    cell.font = boldFont
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(left=noBorder, top=thickBorder, right=thickBorder, bottom=thinBorder)
    row += 1

    # Start of Current Assets (for now, it will also add data for all Non-Current Assets)
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greenFill
    cell.border = Border(left=noBorder, top=thinBorder, right=thickBorder, bottom=noBorder) 
    row += 1

    PPE_tag_info = None
    for asset_tag_info in assets_info:
        # Add formatting
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = greenFill
        cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
        
        # Don't add formula for Non-Current Assets cell
        if asset_tag_info["Tag"] == "us-gaap:AssetsCurrent":
            row += 1
            continue
        # Skip cell if cell is Total Non-Current Assets or Total Assets
        elif asset_tag_info["Tag"] == "us-gaap:Assets" or asset_tag_info["Tag"] == "us-gaap:AssetsNoncurrent":
            continue
        # Get the PPE data if PPE
        elif is_PPE(asset_tag_info["Tag"]):
            PPE_tag_info = asset_tag_info
            continue
        else:
            wb_NAV.cell(row=row, column=col, value=f"={letters[col-2]}{row}*{letters[col-1]}{row}")
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = greenFill
        cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
        row += 1 

    for _ in range(row, asset_row.end):
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = greenFill
        # SG&A Average * Adjustment or PPE * Adjustment
        if row == asset_row.end-3 or (row == asset_row.PPE+1 and PPE_tag_info is not None):
            wb_NAV.cell(row=row, column=col, value=f"={letters[col-2]}{row}*{letters[col-1]}{row}")
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
        if row == asset_row.end-2:
            wb_NAV.cell(row=row, column=col, value=f"=sum({letters[col]}{asset_row.current_asset}:{letters[col]}{asset_row.end-3})")
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            cell.border = Border(left=noBorder, top=thinBorder, right=thickBorder, bottom=noBorder)
        else:
            cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        row += 1

def NAV_liability_adjusted_data(wb_NAV, col, liabilities_info, liability_row):
    wb_NAV.column_dimensions[letters[col]].width = 15
    row = liability_row.start

    # Adjusted
    wb_NAV.cell(row=row, column=col, value="Value")
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greenerFill
    cell.font = boldFont
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(left=noBorder, top=thinBorder, right=thickBorder, bottom=thinBorder)
    row += 1

    # Start of Current Liability (for now, it will also add data for all Non-Current Liability)
    cell = wb_NAV[f"{letters[col]}{row}"]
    cell.fill = greenFill
    cell.border = Border(left=noBorder, top=thinBorder, right=thickBorder, bottom=noBorder) 
    row += 1

    for liability_tag_info in liabilities_info:
        # Skip if Total Non-Current Liabilities or Total Liabilities
        if liability_tag_info["Tag"] == "us-gaap:Liabilities" or liability_tag_info["Tag"] == "us-gaap:LiabilitiesNoncurrent":
            continue

        # Don't add data for Total Current Laibilities
        if liability_tag_info["Tag"] == "us-gaap:LiabilitiesCurrent":
            wb_NAV.cell(row=row, column=col, value=None)
        else:
            wb_NAV.cell(row=row, column=col, value=f"={letters[col-2]}{row}*{letters[col-1]}{row}")
        cell = wb_NAV[f"{letters[col]}{row}"]
        cell.fill = greenFill
        cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
        row += 1 

    for _ in range(row, liability_row.end-1):
        cell = wb_NAV[f"{letters[col]}{row}"]

        cell.fill = greenFill
        # Estimated Option Liability
        if row == liability_row.end-3:
            wb_NAV.cell(row=row, column=col, value=f"={letters[col-2]}{row}*{letters[col-1]}{row}")
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
        # Total Liabilities
        if row == liability_row.end-2:
            wb_NAV.cell(row=row, column=col, value=f"=sum({letters[col]}{liability_row.start+2}:{letters[col]}{liability_row.end-3})")
            cell.number_format = CUSTOM_FORMAT_CURRENCY_ONE
            cell.border = Border(left=noBorder, top=thinBorder, right=thickBorder, bottom=thinBorder)
        else:
            cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        row += 1
    
def NAV_summary_filler(wb_NAV, col, start_row, end_row):
    row = start_row
    for _ in range(start_row, end_row+1):
        cell = wb_NAV[f"{letters[col]}{row}"]
        if row == start_row:
            cell.border = Border(left=noBorder, top=thinBorder, right=thickBorder, bottom=noBorder)
        elif row == end_row:
            cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=thickBorder)
        else:
            cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=noBorder)
        row += 1

def find_which_quarter(all_dates, k_dates):
    i_quarter_away = 0  # variable that lets us know how many quarters to 10-K
    if all_dates[0] == k_dates[0]:
        return i_quarter_away
    
    i_quarter_away += 1
    # Loop through until we reach the 10-K date
    for date in all_dates:
        if date in all_dates:
            # If the current date is a 10-K date, we calculate which quarter we are currently at
            if date in k_dates:
                return 5 - i_quarter_away
        i_quarter_away += 1

    print("Interesting... either we only have 10-Q downloaded or the 10-K and all dates (10-K and 10-Q) dates aren't matching.")
    return 100      # Returning some absurd number for now

def fill_NAV(wb_NAV, assets_info, liabilities_info, shares_outstanding, all_dates, k_dates, SGA_list):
    years = [date.split('-')[0] for date in all_dates]
    wb_NAV.column_dimensions['B'].width = 2
    asset_row = AssetRow()
    liability_row = LiabilityRow()
    summary_row = SummaryRow()
    NAV_assets_titiles(wb_NAV, assets_info, all_dates, k_dates, asset_row)
    NAV_liabilities_titiles(wb_NAV, liabilities_info, asset_row.end, liability_row)
    NAV_summary_titles(wb_NAV, liability_row.end, summary_row)

    # If quarterly, find out if starting point is Q1, Q2, Q3, or FY
    i_quarter = find_which_quarter(all_dates, k_dates)

    iSGA = 0
    for i, date in enumerate(all_dates):
        col = i*3+3
        shares = shares_outstanding[date] if date in shares_outstanding else 0

        if i_quarter == 4:
            i_quarter = 0

        NAV_assets_data(wb_NAV, assets_info, col, asset_row, date, all_dates, k_dates, iSGA, len(k_dates), i_quarter, SGA_list)
        i_quarter += 1
        if date in k_dates:
            iSGA += 1

        NAV_liabilities_data(wb_NAV, liabilities_info, col, liability_row, date)
        NAV_summary_data(wb_NAV, col, asset_row.total_asset, summary_row.start, summary_row.end, shares)

        #Report date
        col += 1
        wb_NAV.cell(row=1, column=col, value=f"{date}")
        cell = wb_NAV[f"{letters[col]}1"]
        cell.alignment = Alignment(horizontal="center")
        NAV_asset_adjustment(wb_NAV, col, assets_info, asset_row, i == 0)
        NAV_liability_adjustment(wb_NAV, col, liabilities_info, liability_row, i == 0)
        cell = wb_NAV[f"{letters[col]}{summary_row.end}"]
        cell.border = Border(left=noBorder, top=noBorder, right=noBorder, bottom=thickBorder)
        col += 1
        NAV_asset_adjusted_data(wb_NAV, col, assets_info, asset_row)
        NAV_liability_adjusted_data(wb_NAV, col, liabilities_info, liability_row)

        NAV_summary_filler(wb_NAV, col, summary_row.start-1, summary_row.end)

        #cell = wb_NAV[f"{letters[col]}{summary_row.end}"]
        #cell.border = Border(left=noBorder, top=noBorder, right=noBorder, bottom=thickBorder)

        # Add some extra formatting if it is the final quarter/year
        if i == len(all_dates)-1:
            for j in range(5):
                cell = wb_NAV[f"{letters[col]}{summary_row.end-j}"]
                if j == 0:
                    cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=thickBorder)
                else:
                    cell.border = Border(left=noBorder, top=noBorder, right=thickBorder, bottom=noBorder)

    return summary_row, [letters[col-2],summary_row.NAV_price], i_quarter