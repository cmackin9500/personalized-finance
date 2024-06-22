import os
from bs4 import BeautifulSoup
import sys
from dataclasses import dataclass
from typing import List
import pandas as pd
import numpy as np
import json
import openpyxl as xl
from datetime import date
import re

from files import read_forms_from_dir, find_latest_form_dir, find_all_form_dir
from xbrl_parse import get_defenition_URI, statement_URI
from edgar_retrieve import get_company_CIK, get_forms_of_type_xbrl, save_all_facts, save_all_forms, download_forms
from html_parse import html_to_facts
from html_process import derived_fs_table, assign_HTMLFact_to_XBRLNode
from util.write_to_csv import *
from util.file_management import read_file
from excel_model.epv import *
from excel_model.cover import *
from excel_model.gv import *
from excel_model.wacc import *
from excel_model.nav import *

import sys
mypath = "./Arelle"
sys.path.insert(0, mypath)

from arelle.CntlrCmdLine import parseAndRun
#from Arelle.arelleCmdLine import runMain

def get_parsing_directories(ticker, parsing_method):
	directory_cfiles = []
	directory_cfiles_10Q = sorted(find_all_form_dir(ticker,"10-Q"))
	directory_cfiles_10Q = list(filter((".DS_Store").__ne__, directory_cfiles_10Q))
	directory_cfiles_10K = sorted(find_all_form_dir(ticker,"10-K"))
	directory_cfiles_10K = list(filter((".DS_Store").__ne__, directory_cfiles_10K))

	# Parse every single quarter and 10-K
	if parsing_method == 'q':
		directory_cfiles = sorted(directory_cfiles_10Q + directory_cfiles_10K)

	# Parse all 10-Q past the most recent 10-K
	elif parsing_method == 'r':
		directory_cfiles = directory_cfiles_10K + []	# Just adding + [] so it will not be a reference to directory_cfiles_10K
		for i, cfile_10Q in enumerate(directory_cfiles_10Q):
			if cfile_10Q > directory_cfiles_10K[-1]:
				directory_cfiles += directory_cfiles_10Q[i:]
				break
	# Parse the latest only
	# TODO: this will work once epv_info shit is sorted
	elif parsing_method == 'l':
		if directory_cfiles_10K[-1] > directory_cfiles_10Q[-1]:
			directory_cfiles = directory_cfiles_10K[-1:]
		else:
			directory_cfiles = directory_cfiles_10Q[-1:]
	# Only parse 10-K
	else:
		directory_cfiles = directory_cfiles_10K

	return directory_cfiles, directory_cfiles_10K, directory_cfiles_10Q
	

def convert_data_to_fs_dict(curConcept, fs_data, fs_dict, visited, depth):
	if curConcept is None or curConcept in visited:
		return
	
	relationshipSet = fs_data["relationship"]
	modelRelationshipsFrom = relationshipSet.modelRelationshipsFrom
	visited.add(curConcept)
	tag = curConcept.vQname().prefix + ':' + curConcept.name
	if tag not in fs_dict:
		fs_dict[tag] = {}
		fs_dict[tag]["facts"] = {}
		fs_dict[tag]["label"] = curConcept.label()
		fs_dict[tag]["depth"] = depth
	
	if curConcept.qname in fs_data["facts"]:
		for date in fs_data["facts"][curConcept.qname]:
			val = fs_data["facts"][curConcept.qname][date][0]
			fs_dict[tag]["facts"][date] = val

	conceptsToAdd = []
	modelRelationshipsFromList = modelRelationshipsFrom[curConcept]
	for cur_modelRelationshipsFrom in modelRelationshipsFromList:
		toConcept = cur_modelRelationshipsFrom.toModelObject
		# Only add the child concept if it has not been added to the parent yet
		# TODO: This could pose an issue in the future when we introduce duplicate tags
		if toConcept not in visited:
			conceptsToAdd.append(toConcept)		
	
	for concept in conceptsToAdd:
		convert_data_to_fs_dict(concept, fs_data, fs_dict, visited, depth+1)

	return

def fs_list_insert_or_add_concept(fs_list, cur_year_fs_dict, bRemoveDuplicateDates=False):
	if fs_list == []:
		for conceptName in cur_year_fs_dict:
			fs_list.append(
				{"tag": conceptName, 
	 			"label": cur_year_fs_dict[conceptName]["label"],
				"facts": cur_year_fs_dict[conceptName]["facts"], 
				"depth":  cur_year_fs_dict[conceptName]["depth"]})
	else:

		if bRemoveDuplicateDates:
			# Remove exisiting dates
			new_dates = set()
			for conceptName in cur_year_fs_dict:
				for date in cur_year_fs_dict[conceptName]["facts"]:
					new_dates.add(date)
			for tag_info in fs_list:
				for new_date in new_dates:
					if new_date in tag_info["facts"]:
						del tag_info["facts"][new_date]

		index = 0
		for conceptName in cur_year_fs_dict:
			for i in range(len(fs_list)):
				next = False
				if conceptName == fs_list[i]["tag"]:
					fs_list[i]["facts"] = fs_list[i]["facts"] | cur_year_fs_dict[conceptName]["facts"]
					index = i+1
					next = True
					break
			if next: continue
			fs_list.insert(index, 
				{	
					"tag": conceptName, 
	   				"facts": cur_year_fs_dict[conceptName]["facts"],
					"depth": cur_year_fs_dict[conceptName]["depth"]
				})
	return fs_list

def reorder_and_add_blank_columns(df_fs):
	df_fs = df_fs.reindex(sorted(df_fs.columns), axis=1)
	df_fs = df_fs.iloc[:, ::-1]
	[df_fs.insert(i, '', np.nan, allow_duplicates=True) for i in range(df_fs.shape[1], 1, -1)]
	return df_fs

def fill_ratio_tags(ratio_tags, all_fs_info, bMain=False):
	all_ratio_tags = []
	# Loop through each financial statement concepts
	for fs in ratio_tags:
		fs_info = all_fs_info[fs]
		# Loop through the concepts
		for concept in ratio_tags[fs]:
			if not bMain:
				if concept in ["Goodwill", "Intangible Assets", "SG&A", "CF from Investing Portion"]: continue
			cur_concept = {"concept": concept}
			for tag_info in fs_info:
				tag = tag_info['tag'].split(':')[-1]
				if tag in ratio_tags[fs][concept]:
					for key in tag_info:
						if key == "label" or key == "tag": continue
						val = tag_info[key]
						if concept == "CAPEX" or concept == "Dividend":
							val = -abs(val)
						if key in cur_concept:
							cur_concept[key] += val
						else:
							cur_concept[key] = val
			all_ratio_tags.append(cur_concept)
	return all_ratio_tags

def get_epv_info_from_fs(epv_info, epv_tags, fs_list, cur_year):
	for tag_info in fs_list:
		if ':' in tag_info['tag']:
			tag = tag_info['tag'].split(":")[1]
		else:
			tag = tag_info['tag']

		for key in epv_tags:
			epv_tag = epv_tags[key]
			if tag in epv_tag:
				if key in epv_info[cur_year] and cur_year in tag_info:
					epv_info[cur_year][key] += tag_info[cur_year]
				else:
					if cur_year not in tag_info:
						continue
					epv_info[cur_year][key] = tag_info[cur_year]


def get_SGA_values(SGA_tags, all_is_info, directory_cfiles_10Q):
	# Get the SG&A values
	SGA_values = {}
	for tag_info in all_is_info:
		if tag_info["tag"].split(':')[1] in SGA_tags:
			for key in tag_info:
				if key == "tag" or key == "label" or key in directory_cfiles_10Q:
					continue
				if key in SGA_values:
					SGA_values[key] += tag_info[key]
				else:
					SGA_values[key] = tag_info[key]
	return SGA_values

def remove_usgaap(label):
	if label.startswith("us-gaap:"):
		label = label.split(':')[-1]
		label = re.sub(r'((?<=[a-z])[A-Z]|(?<!\A)[A-Z](?=[a-z]))', r' \1', label)
	elif label.startswith("us-gaap_"):
		label = label.split('_')[-1]
		label = re.sub(r'((?<=[a-z])[A-Z]|(?<!\A)[A-Z](?=[a-z]))', r' \1', label)
	return label
	
def run_arelle():
	ticker = input("Enter ticker: ")
	scale = input("Enter scale: ")
	industry = input("Enter indsutry: ")
	parsing_method = input("Enter parsing method: ")
	
	div = 1000
	if scale == 'm':
		div = 1000000
	elif scale == '':
		div = 1

	#download_forms(ticker, False, False)
	download_forms(ticker, False, False)
	directory_cfiles, directory_cfiles_10K, directory_cfiles_10Q = get_parsing_directories(ticker, parsing_method)
	
	fs_data = {}
	fs_list = {'bs': [], 'is': [], 'cf': []}
	for cur_year in directory_cfiles:
		try:
			form_type = "10-K" if cur_year in directory_cfiles_10K else "10-Q"
			cfiles = read_forms_from_dir(f"forms/{ticker}/{form_type}/{cur_year}")
			zip_file = "/Users/caseymackinnon/Desktop/Personalized Finance/edgar/" + cfiles.zip
			#zip_file = "./" + cfiles.zip
			arg = ['-f', zip_file, "--factTable=test.txt"]
			data = parseAndRun(arg)


			statement_roleURI = get_defenition_URI(cfiles.xsd, 'Statement')
			roleURIs = [roleURI for roleURI in statement_roleURI]
			fs_data[cur_year] = {}
			for fs in ('bs', 'is', 'cf'):
				fs_roleURI = statement_URI(roleURIs, fs)
				fs_data[cur_year][fs] = data[fs_roleURI]
				
				cur_year_fs_data = fs_data[cur_year][fs]
				cur_year_fs_dict, visited = dict(), set()
				rootConcept = cur_year_fs_data["relationship"].rootConcepts[0]
				convert_data_to_fs_dict(rootConcept, cur_year_fs_data, cur_year_fs_dict, visited, 0)

				fs_list[fs] = fs_list_insert_or_add_concept(fs_list[fs], cur_year_fs_dict)

		except:
			print(f"Failed to prase for {cur_year}.")
	
	all_fs_info = {'bs': [], 'is': [], 'cf': []}
	for fs in fs_list:
		cur_fs_list = fs_list[fs]

		for tag_info in cur_fs_list:
			dict_to_add = {}
			dict_to_add['tag'] = tag_info['tag']
			if 'label' in tag_info:
				label = tag_info['label']
				label = remove_usgaap(label)
				dict_to_add['label'] = label
			else:
				label = remove_usgaap(tag_info['tag'])
				label = re.sub(r'((?<=[a-z])[A-Z]|(?<!\A)[A-Z](?=[a-z]))', r' \1', label)
				dict_to_add['label'] = label
			for date in tag_info['facts']:
				str_date = date.strftime('%Y-%m-%d')
				str_val = tag_info['facts'][date]
				if str_val != '(nil)':
					val = float(str_val.replace(',',''))/div
					dict_to_add[str_date] = val
				else:
					dict_to_add[str_date] = tag_info['facts'][date]

			all_fs_info[fs].append(dict_to_add)

	with open('./tags/ratio_tags.json') as f:
		file_contents = f.read()
		ratio_tags = json.loads(file_contents)
	
	all_ratio_tags = fill_ratio_tags(ratio_tags, all_fs_info, True)	
	all_other_ratio_tags = fill_ratio_tags(ratio_tags, all_fs_info)			
			
	writer = pd.ExcelWriter(f"./excel/{ticker}.xlsx", engine='xlsxwriter')
	df_bs = pd.DataFrame(all_fs_info['bs'])
	df_bs = reorder_and_add_blank_columns(df_bs)

	df_is = pd.DataFrame(all_fs_info['is'])
	df_is = reorder_and_add_blank_columns(df_is)

	df_cf = pd.DataFrame(all_fs_info['cf'])
	df_cf = reorder_and_add_blank_columns(df_cf)

	df_concepts = pd.DataFrame(all_ratio_tags)
	df_concepts = reorder_and_add_blank_columns(df_concepts)

	df_other_concepts = pd.DataFrame(all_other_ratio_tags)
	df_other_concepts = reorder_and_add_blank_columns(df_other_concepts)

	df_bs.to_excel(writer, sheet_name='Balance Sheet')
	df_is.to_excel(writer, sheet_name='Income Statement')
	df_cf.to_excel(writer, sheet_name='Cash Flow')
	df_concepts.to_excel(writer, sheet_name='Ratios')
	df_other_concepts.to_excel(writer, sheet_name='Other Ratios')
	writer.save()



	with open('./tags/facts_tags.json') as f:
		file_contents = f.read()
		SGA_tags = json.loads(file_contents)["SG&A"]
	SGA_values = get_SGA_values(SGA_tags, all_fs_info['is'], directory_cfiles_10Q)
	SGA_list = list(SGA_values.values())

	epv_info = {}
	epv_tags = json.loads(read_file(f"./tags/epv_{industry}_tags.json"))

	dates = set()
	for fs in ('bs', 'is', 'cf'):
		cur_fs_list = fs_list[fs]
		for tag_info in cur_fs_list:
			for date in tag_info["facts"]:
				str_date = date.strftime('%Y-%m-%d')
				dates.add(str_date)

	for date in dates:
		epv_info[date] = {}
		for fs in ('bs', 'is', 'cf'):
			cur_fs_info = all_fs_info[fs]
			get_epv_info_from_fs(epv_info, epv_tags, cur_fs_info, date)
	
	epv_info = dict(sorted(epv_info.items()))
	
	wb = xl.Workbook()
	wb_cover = wb["Sheet"]
	wb_cover.title = "COVER"
	fill_cover(wb_cover, ticker, 5)

	wb.create_sheet("WACC")
	wb_WACC = wb["WACC"]
	fill_wacc(wb_WACC, scale)
	wb.create_sheet("NAV")
	wb_NAV = wb["NAV"]

	shares_outsanding = {}
	iAsset, iLiabilities = [None, None] , [None, None]
	for i, info in enumerate(all_fs_info['bs']):
		if info["tag"] == "us-gaap:AssetsAbstract":
			iAsset[0] = i
			continue
		if info["tag"] == "us-gaap:LiabilitiesAndStockholdersEquityAbstract":
			iAsset[1] = i-1
			iLiabilities[0] = i
		if info["tag"] == "us-gaap:Liabilities":
			iLiabilities[1] = i
			continue
		if iLiabilities is not None and info["tag"] == "us-gaap:PreferredStockValue" or info["tag"] == "us-gaap:CommitmentsAndContingencies" or info["tag"] == "us-gaap:CommonStockValue":
			iLiabilities[1] = i-1
			continue
	if iAsset is None:
		print("iAsset is not found")
	if iLiabilities is None:
		print("iLiabilities is not found")

	assets_info = all_fs_info['bs'][iAsset[0]: iAsset[1]]
	liabilities_info = all_fs_info['bs'][iLiabilities[0]: iLiabilities[1]]
	NAV_summary_row, iNAVPriceCoord, i_quarter = fill_NAV(wb_NAV, assets_info, liabilities_info, shares_outsanding, directory_cfiles, directory_cfiles_10K, SGA_list)

	wb.create_sheet("EPV")
	wb_EPV = wb["EPV"]
	EPV_rows, iEPVPriceCoord = fill_epv(wb_EPV, industry, epv_info, NAV_summary_row.shares)

	iNAVRow = NAV_summary_row.shares-1
	wb.create_sheet("GV")
	wb_GV = wb["GV"]
	years = list(dates)
	iGVPriceCoord = fill_gv(wb_GV, years, iNAVRow, EPV_rows)

	# Some hard coded stuff
	wb_cover.cell(row=2, column=2, value=ticker)
	wb_cover.cell(row=9, column=3, value=f"=NAV!{iNAVPriceCoord[0]+str(iNAVPriceCoord[1])}")
	wb_cover.cell(row=10, column=3, value=f"=EPV!{iEPVPriceCoord[0]+str(iEPVPriceCoord[1])}")
	wb_cover.cell(row=11, column=3, value=f"=GV!{iGVPriceCoord[0]+str(iGVPriceCoord[1])}")

	sBasis = "FY"
	if parsing_method == 'q':
		sBasis = "Q"

	year = int(directory_cfiles_10K[-1].split('-')[0])
	i_quarter -= 1
	if i_quarter > 0 and i_quarter < 4:
		year += 1
		sBasis = "Q"
	else:
		sBasis = "FY"
		i_quarter = ''
	#return wb
	wb.save(f"/Users/caseymackinnon/Desktop/Personalized Finance/edgar/excel/{ticker} - {sBasis}{i_quarter} {year} - {directory_cfiles[-1]} TEST.xlsx")

if __name__ == "__main__":
    run_arelle()